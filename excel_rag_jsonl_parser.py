#!/usr/bin/env python3
"""
Excel RAG JSONL Parser

목적
- 복잡한 엑셀 표를 RAG 적재용 JSONL로 변환한다.
- 병합 셀, 다단 헤더, 계층형 행, 전결권자 매트릭스, 합의/수신 메타데이터를 보존한다.
- 임베딩에는 JSON 원문이 아니라 content_text 필드를 사용하도록 설계한다.

주의
- 이 스크립트는 전결규정표를 우선 지원하지만, 구조 감지/청크 스키마는 범용 엑셀 RAG 적재용으로 확장 가능하게 작성되어 있다.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

MARKER_CHARS = {"○", "●", "◎", "◯", "O", "o"}
HEADER_ITEM_TERMS = {"전 결 사 항", "전결사항", "전결 사항"}
HEADER_IGNORE_TERMS = {"전결권자", "전 결 사 항", "전결사항", "전결 사항"}
SPECIAL_HEADER_TERMS = {"합의", "수신", "참조", "비고"}
NOTE_PREFIXES = ("※", "*", "주)", "주:", "단,", "단 ", "(단,", "(단 ", "(※", "- ")


def clean_text(value: Any) -> str:
    """셀 값을 검색/비교 가능한 문자열로 정규화한다."""
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r" *\n *", "\n", text)
    return text.strip()


def one_line(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).strip()


def is_marker(value: Any) -> bool:
    text = one_line(value)
    if not text:
        return False
    compact = text.replace(" ", "")
    if compact in MARKER_CHARS:
        return True
    return bool(re.fullmatch(r"[○●◎◯Oo]+", compact))


def is_note_text(text: str) -> bool:
    t = one_line(text)
    return any(t.startswith(prefix) for prefix in NOTE_PREFIXES)


def split_keywords(text: str) -> List[str]:
    """간단한 한국어/영문 키워드 추출. 형태소 분석기 없이 안전하게 동작하도록 보수적으로 구성."""
    text = one_line(text)
    tokens = re.split(r"[\s,;/·>\(\)\[\]\{\}:]+", text)
    out: List[str] = []
    seen = set()
    for token in tokens:
        token = token.strip(" .-–—①②③④⑤⑥⑦⑧⑨⑩")
        if len(token) < 2:
            continue
        if token in seen:
            continue
        seen.add(token)
        out.append(token)
    return out[:40]


def stable_id(*parts: Any) -> str:
    raw = "::".join(one_line(p) for p in parts)
    digest = hashlib.md5(raw.encode("utf-8")).hexdigest()[:10]
    safe_prefix = re.sub(r"[^0-9A-Za-z가-힣_.-]+", "_", one_line(parts[0]))[:40]
    return f"{safe_prefix}::{digest}"


@dataclass
class RegionSpec:
    min_row: int
    min_col: int
    max_row: int
    max_col: int

    @property
    def address(self) -> str:
        return f"{get_column_letter(self.min_col)}{self.min_row}:{get_column_letter(self.max_col)}{self.max_row}"


@dataclass
class SheetLayout:
    sheet_name: str
    title: str
    region: RegionSpec
    header_top: int
    data_start: int
    item_cols: List[int]
    approver_cols: Dict[int, str]
    special_cols: Dict[int, str]


class MergedCellResolver:
    """병합 셀의 logical value를 좌상단 셀 값으로 복원한다."""

    def __init__(self, ws):
        self.ws = ws
        self._origin: Dict[Tuple[int, int], Tuple[int, int, str]] = {}
        for merged_range in ws.merged_cells.ranges:
            coord = str(merged_range)
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    self._origin[(r, c)] = (merged_range.min_row, merged_range.min_col, coord)

    def raw(self, row: int, col: int) -> str:
        return clean_text(self.ws.cell(row, col).value)

    def value(self, row: int, col: int) -> str:
        if (row, col) in self._origin:
            origin_row, origin_col, _ = self._origin[(row, col)]
            return clean_text(self.ws.cell(origin_row, origin_col).value)
        return clean_text(self.ws.cell(row, col).value)

    def merged_range(self, row: int, col: int) -> Optional[str]:
        origin = self._origin.get((row, col))
        return origin[2] if origin else None

    def merged_bounds_containing_text(self, text_terms: Iterable[str], max_scan_row: int = 30) -> Optional[Tuple[int, int, int, int]]:
        terms = {one_line(t).replace(" ", "") for t in text_terms}
        # 먼저 병합 범위의 좌상단 값을 확인한다.
        for mr in self.ws.merged_cells.ranges:
            if mr.min_row > max_scan_row:
                continue
            value = one_line(self.ws.cell(mr.min_row, mr.min_col).value).replace(" ", "")
            if value in terms:
                return (mr.min_row, mr.min_col, mr.max_row, mr.max_col)
        # 병합이 아니라 단일 셀 헤더인 경우
        for r in range(1, min(self.ws.max_row, max_scan_row) + 1):
            for c in range(1, self.ws.max_column + 1):
                value = one_line(self.ws.cell(r, c).value).replace(" ", "")
                if value in terms:
                    return (r, c, r, c)
        return None


def extract_title(ws) -> str:
    candidates = []
    for r in range(1, min(ws.max_row, 6) + 1):
        for c in range(1, min(ws.max_column, 15) + 1):
            text = one_line(ws.cell(r, c).value)
            if not text:
                continue
            if "기준표" in text or "위임전결" in text:
                candidates.append(text)
    return candidates[0] if candidates else ws.title


def auto_filter_region(ws) -> Optional[RegionSpec]:
    ref = getattr(ws.auto_filter, "ref", None)
    if not ref:
        return None
    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        return RegionSpec(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col)
    except Exception:
        return None


def row_has_any(ws, row: int, min_col: int, max_col: int) -> bool:
    return any(clean_text(ws.cell(row, c).value) for c in range(min_col, max_col + 1))


def detect_region(ws, resolver: MergedCellResolver) -> Optional[RegionSpec]:
    # 자동 필터가 잡힌 표는 그 영역을 우선 신뢰한다.
    autofilter = auto_filter_region(ws)
    if autofilter:
        return autofilter

    header_bounds = resolver.merged_bounds_containing_text(HEADER_ITEM_TERMS)
    if not header_bounds:
        return None
    header_top, item_min_col, _, item_max_col = header_bounds

    # 헤더 주변에서 오른쪽 끝 컬럼 추정
    max_col = item_max_col
    for c in range(item_min_col, min(ws.max_column, item_max_col + 15) + 1):
        values = [one_line(resolver.value(r, c)) for r in range(header_top, min(header_top + 3, ws.max_row) + 1)]
        if any(v for v in values):
            max_col = c

    # 긴 공백 구간이 나오면 표가 끝났다고 판단한다.
    last_data_row = header_top
    blank_run = 0
    for r in range(header_top + 1, ws.max_row + 1):
        has = row_has_any(ws, r, item_min_col, max_col)
        if has:
            last_data_row = r
            blank_run = 0
        else:
            blank_run += 1
            if blank_run >= 30 and last_data_row > header_top:
                break
    return RegionSpec(min_row=header_top, min_col=item_min_col, max_row=last_data_row, max_col=max_col)


def detect_layout(ws, resolver: MergedCellResolver) -> Optional[SheetLayout]:
    region = detect_region(ws, resolver)
    if not region:
        return None

    item_bounds = resolver.merged_bounds_containing_text(HEADER_ITEM_TERMS)
    if not item_bounds:
        return None
    item_header_top, item_min_col, item_header_bottom, item_max_col = item_bounds
    item_cols = list(range(item_min_col, item_max_col + 1))

    # data_start: 헤더 하단 이후 처음으로 item column에 실제 내용이 등장하는 행
    data_start = item_header_bottom + 1
    for r in range(item_header_bottom + 1, min(region.max_row, item_header_bottom + 10) + 1):
        item_values = [one_line(resolver.raw(r, c) or resolver.value(r, c)) for c in item_cols]
        joined = " ".join(item_values)
        if joined and not any(term.replace(" ", "") in joined.replace(" ", "") for term in HEADER_ITEM_TERMS):
            data_start = r
            break

    approver_cols: Dict[int, str] = {}
    special_cols: Dict[int, str] = {}
    header_scan_rows = range(item_header_top, max(data_start, item_header_top + 1))

    for c in range(item_max_col + 1, region.max_col + 1):
        header_values = [one_line(resolver.value(r, c)) for r in header_scan_rows]
        unique_values = []
        seen = set()
        for value in header_values:
            if not value or value in seen:
                continue
            seen.add(value)
            unique_values.append(value)

        # 합의/수신 등 특수 메타데이터 컬럼
        matched_special = None
        for value in unique_values:
            compact = value.replace(" ", "")
            for term in SPECIAL_HEADER_TERMS:
                if compact == term.replace(" ", ""):
                    matched_special = term
                    break
            if matched_special:
                break
        if matched_special:
            special_cols[c] = matched_special
            continue

        # 전결권자 상위 헤더는 제외하고 실제 하위 라벨을 찾는다.
        candidates = [
            v
            for v in unique_values
            if v.replace(" ", "") not in {term.replace(" ", "") for term in HEADER_IGNORE_TERMS}
            and v not in SPECIAL_HEADER_TERMS
        ]
        if candidates:
            approver_cols[c] = candidates[-1]

    return SheetLayout(
        sheet_name=ws.title,
        title=extract_title(ws),
        region=region,
        header_top=item_header_top,
        data_start=data_start,
        item_cols=item_cols,
        approver_cols=approver_cols,
        special_cols=special_cols,
    )


def infer_hierarchy_level(text: str, fallback_level: int = 0) -> int:
    """항목 번호 패턴으로 계층 레벨을 추정한다."""
    t = one_line(text)
    if not t:
        return fallback_level
    if re.match(r"^[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+\.", t):
        return 0
    if re.match(r"^\d+\.", t):
        return 0
    if re.match(r"^[가-힣]\.", t):
        return 1
    if re.match(r"^\(\d+\)", t):
        return 2
    if re.match(r"^\([가-힣]\)", t):
        return 3
    if re.match(r"^[①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮]", t):
        return 4
    if is_note_text(t):
        return fallback_level
    return fallback_level


def extract_item(ws, resolver: MergedCellResolver, row: int, item_cols: List[int]) -> Tuple[str, Optional[int], bool]:
    """행에서 항목 텍스트를 추출한다. 실제 입력 셀(raw)을 우선하고, 필요한 경우 logical value를 보조로 사용한다."""
    for idx, col in enumerate(item_cols):
        raw = one_line(resolver.raw(row, col))
        if raw:
            return raw, col, False
    # merged value가 세로로 이어져 있고 현재 행에 marker가 있는 경우를 대비한 보조 처리
    for idx, col in enumerate(item_cols):
        logical = one_line(resolver.value(row, col))
        if logical and logical.replace(" ", "") not in {term.replace(" ", "") for term in HEADER_ITEM_TERMS}:
            return logical, col, True
    return "", None, False


def parse_code_map(wb) -> Dict[str, str]:
    """Index 시트의 약어 매핑을 보수적으로 추출한다."""
    if "Index" not in wb.sheetnames:
        return {}
    ws = wb["Index"]
    mapping: Dict[str, str] = {}
    # 일반적으로 B:C, E:F가 약어/설명 쌍이다.
    for left_col, right_col in [(2, 3), (5, 6)]:
        for r in range(1, ws.max_row + 1):
            key = one_line(ws.cell(r, left_col).value)
            value = one_line(ws.cell(r, right_col).value)
            if not key or not value:
                continue
            if key == "▼" or key.startswith("*"):
                continue
            # 표 본문이 섞인 구간을 약어로 오인하지 않도록 제한
            if len(key) > 30 or key.startswith("(") or key.startswith("①"):
                continue
            if re.search(r"[가-힣A-Za-z]", key) and key not in mapping:
                mapping[key] = value
    return mapping


def expand_org_codes(raw: str, code_map: Dict[str, str]) -> List[Dict[str, str]]:
    """합의/수신 값의 약어를 원문과 확장명으로 함께 보존한다."""
    raw = one_line(raw)
    if not raw:
        return []
    # 쉼표 기준 + 괄호 안 코드도 별도 인식
    candidates = []
    for part in re.split(r"[,，/]", raw):
        part = part.strip()
        if not part:
            continue
        candidates.append(part)
        for inner in re.findall(r"\(([^)]+)\)", part):
            candidates.append(inner.strip())

    result = []
    seen = set()
    for token in candidates:
        token_clean = token.strip().strip("()")
        if not token_clean or token_clean in seen:
            continue
        seen.add(token_clean)
        result.append({"raw": token_clean, "expanded": code_map.get(token_clean, "")})
    return result


def compose_content_text(title: str, sheet: str, path: List[str], approvers: List[str], special_values: Dict[str, str], chunk_type: str) -> str:
    path_text = " > ".join(path)
    base = f"{title}의 {sheet} 시트에서 '{path_text}' 항목"
    if chunk_type == "hierarchy_node":
        return f"{base}은/는 하위 전결사항을 포함하는 상위 항목이다."
    if chunk_type == "note":
        return f"{base}에 대한 참고/주의사항이다."
    parts = []
    if approvers:
        parts.append(f"전결권자는 {', '.join(approvers)}이다")
    for key, value in special_values.items():
        if value:
            parts.append(f"{key}는 {value}이다")
    if parts:
        return f"{base}의 " + ", ".join(parts) + "."
    return f"{base}에 대한 전결 기준 항목이다."


def make_source(source_file: str, sheet: str, row: int, layout: SheetLayout) -> Dict[str, Any]:
    return {
        "file": source_file,
        "sheet": sheet,
        "row": row,
        "range": f"{get_column_letter(min(layout.item_cols))}{row}:{get_column_letter(layout.region.max_col)}{row}",
    }


def parse_delegation_sheet(ws, resolver: MergedCellResolver, layout: SheetLayout, source_file: str, code_map: Dict[str, str]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    stack: List[str] = []
    section_stats: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"decision_count": 0, "approvers": Counter(), "children": set()})

    # 표 요약 청크는 본문 파싱 후 통계를 넣기 위해 마지막에 보완한다.
    row_records: List[Dict[str, Any]] = []

    for row in range(layout.data_start, layout.region.max_row + 1):
        item, item_col, came_from_merge = extract_item(ws, resolver, row, layout.item_cols)

        approvers: List[str] = []
        for col, label in layout.approver_cols.items():
            if is_marker(resolver.value(row, col)) or is_marker(resolver.raw(row, col)):
                approvers.append(label)

        special_values: Dict[str, str] = {}
        special_expanded: Dict[str, List[Dict[str, str]]] = {}
        for col, label in layout.special_cols.items():
            value = one_line(resolver.value(row, col) or resolver.raw(row, col))
            # 병합 헤더가 본문 행에 내려온 경우는 제외
            if value and value.replace(" ", "") not in {label.replace(" ", ""), "전결권자"}:
                special_values[label] = value
                special_expanded[label] = expand_org_codes(value, code_map)

        # 완전 공백 행은 스킵
        if not item and not approvers and not special_values:
            continue

        # item이 없는 marker 행은 직전 path의 연장으로 처리한다.
        if not item and (approvers or special_values):
            item = stack[-1] if stack else f"행 {row}"
            item_col = layout.item_cols[-1]
            came_from_merge = True

        if not item:
            continue

        fallback_level = 0
        if item_col in layout.item_cols:
            fallback_level = layout.item_cols.index(item_col)
        level = infer_hierarchy_level(item, fallback_level=fallback_level)
        note = is_note_text(item)

        if note:
            path = stack + [item]
            chunk_type = "note"
        else:
            if level < len(stack):
                stack = stack[:level]
            while len(stack) < level:
                # 레벨 점프가 생기는 경우 빈 부모를 만들지 않고 현재 위치에 붙인다.
                level = len(stack)
            if len(stack) == level:
                stack.append(item)
            else:
                stack[level] = item
            path = stack[: level + 1]
            chunk_type = "delegation_rule" if approvers or special_values else "hierarchy_node"

        is_decision_row = bool(approvers or special_values)
        top_section = path[0] if path else ""
        if top_section:
            section_stats[top_section]["children"].add(path[1] if len(path) > 1 else path[0])
            if is_decision_row:
                section_stats[top_section]["decision_count"] += 1
                for approver in approvers:
                    section_stats[top_section]["approvers"][approver] += 1

        fields = {
            "항목": item,
            "경로": " > ".join(path),
            "전결권자": approvers,
            **special_values,
        }
        facts = []
        for approver in approvers:
            facts.append({"predicate": "전결권자", "value": approver})
        for key, value in special_values.items():
            facts.append({"predicate": key, "value": value, "expanded": special_expanded.get(key, [])})

        content_text = compose_content_text(layout.title, ws.title, path, approvers, special_values, chunk_type)
        keywords = []
        for part in path + approvers + list(special_values.values()):
            keywords.extend(split_keywords(part))
        # 순서 보존 중복 제거
        keywords = list(dict.fromkeys(keywords))[:60]

        record = {
            "id": stable_id(source_file, ws.title, row, chunk_type, " > ".join(path), ",".join(approvers), json.dumps(special_values, ensure_ascii=False)),
            "source_file": source_file,
            "sheet": ws.title,
            "excel_row": row,
            "range": f"{get_column_letter(min(layout.item_cols))}{row}:{get_column_letter(layout.region.max_col)}{row}",
            "chunk_type": chunk_type,
            "region_type": "hierarchical_matrix",
            "title": layout.title,
            "path": path,
            "level": level,
            "item": item,
            "fields": fields,
            "facts": facts,
            "approver": ", ".join(approvers),
            "consultation": special_values.get("합의", ""),
            "receiver": special_values.get("수신", ""),
            "content_text": content_text,
            "search_keywords": keywords,
            "answer_template": "{path} 항목의 전결권자는 {approver}입니다." if approvers else "{path} 항목은 상위 항목 또는 참고사항입니다.",
            "source": make_source(source_file, ws.title, row, layout),
            "metadata": {
                "is_decision_row": is_decision_row,
                "has_approver": bool(approvers),
                "has_special_values": bool(special_values),
                "came_from_merged_cell": came_from_merge,
                "parser_version": "excel-rag-jsonl-parser-v1",
                "confidence": 0.94 if is_decision_row and path else 0.82,
            },
        }
        row_records.append(record)

    # section summary 청크 생성
    section_records = []
    for section, stats in section_stats.items():
        top_approvers = [name for name, _ in stats["approvers"].most_common(5)]
        content = f"{layout.title}의 {ws.title} 시트에서 '{section}' 섹션은 {stats['decision_count']}개의 전결/합의 기준 행을 포함한다."
        if top_approvers:
            content += f" 주요 전결권자는 {', '.join(top_approvers)}이다."
        section_records.append({
            "id": stable_id(source_file, ws.title, "section", section),
            "source_file": source_file,
            "sheet": ws.title,
            "chunk_type": "section_summary",
            "region_type": "hierarchical_matrix",
            "title": layout.title,
            "path": [section],
            "fields": {
                "섹션": section,
                "decision_count": stats["decision_count"],
                "top_approvers": top_approvers,
                "child_count": len(stats["children"]),
            },
            "facts": [],
            "content_text": content,
            "search_keywords": split_keywords(section) + top_approvers,
            "source": {
                "file": source_file,
                "sheet": ws.title,
                "range": layout.region.address,
            },
            "metadata": {
                "parser_version": "excel-rag-jsonl-parser-v1",
                "confidence": 0.86,
            },
        })

    decision_count = sum(1 for r in row_records if r["metadata"]["is_decision_row"])
    hierarchy_count = sum(1 for r in row_records if r["chunk_type"] == "hierarchy_node")
    note_count = sum(1 for r in row_records if r["chunk_type"] == "note")
    table_summary = {
        "id": stable_id(source_file, ws.title, "table_summary"),
        "source_file": source_file,
        "sheet": ws.title,
        "chunk_type": "table_summary",
        "region_type": "hierarchical_matrix",
        "title": layout.title,
        "path": [layout.title],
        "fields": {
            "sheet": ws.title,
            "region": layout.region.address,
            "row_count": len(row_records),
            "decision_count": decision_count,
            "hierarchy_count": hierarchy_count,
            "note_count": note_count,
            "approver_columns": list(layout.approver_cols.values()),
            "special_columns": list(layout.special_cols.values()),
        },
        "facts": [],
        "content_text": f"{layout.title}의 {ws.title} 시트는 전결사항별 전결권자와 합의/수신 등 메타데이터를 계층형 표로 정의한다. 총 {decision_count}개의 전결/합의 기준 행이 파싱되었다.",
        "search_keywords": split_keywords(layout.title + " " + ws.title) + list(layout.approver_cols.values()) + list(layout.special_cols.values()),
        "source": {"file": source_file, "sheet": ws.title, "range": layout.region.address},
        "metadata": {"parser_version": "excel-rag-jsonl-parser-v1", "confidence": 0.9},
    }

    records.extend([table_summary])
    records.extend(section_records)
    records.extend(row_records)
    return records


def parse_index_sheet(wb, source_file: str, code_map: Dict[str, str]) -> List[Dict[str, Any]]:
    if "Index" not in wb.sheetnames:
        return []
    ws = wb["Index"]
    records = []
    for code, meaning in sorted(code_map.items()):
        records.append({
            "id": stable_id(source_file, "Index", code, meaning),
            "source_file": source_file,
            "sheet": "Index",
            "chunk_type": "code_mapping",
            "region_type": "key_value_index",
            "title": "약어 매핑",
            "path": ["약어 매핑", code],
            "fields": {"약어": code, "의미": meaning},
            "facts": [{"predicate": "약어의미", "value": meaning}],
            "content_text": f"위임전결기준표에서 약어 '{code}'은/는 '{meaning}'을 의미한다.",
            "search_keywords": [code, meaning],
            "source": {"file": source_file, "sheet": ws.title},
            "metadata": {"parser_version": "excel-rag-jsonl-parser-v1", "confidence": 0.78},
        })
    return records


def parse_workbook(input_path: Path) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    wb = load_workbook(input_path, data_only=True)
    source_file = input_path.name
    code_map = parse_code_map(wb)
    all_records: List[Dict[str, Any]] = []
    layouts: Dict[str, Any] = {}

    for ws in wb.worksheets:
        if ws.title == "Index":
            continue
        resolver = MergedCellResolver(ws)
        layout = detect_layout(ws, resolver)
        if not layout:
            continue
        records = parse_delegation_sheet(ws, resolver, layout, source_file, code_map)
        all_records.extend(records)
        layouts[ws.title] = {
            "title": layout.title,
            "region": layout.region.address,
            "header_top": layout.header_top,
            "data_start": layout.data_start,
            "item_cols": [get_column_letter(c) for c in layout.item_cols],
            "approver_cols": {get_column_letter(k): v for k, v in layout.approver_cols.items()},
            "special_cols": {get_column_letter(k): v for k, v in layout.special_cols.items()},
            "records": len(records),
            "decision_rows": sum(1 for r in records if r.get("metadata", {}).get("is_decision_row")),
        }

    index_records = parse_index_sheet(wb, source_file, code_map)
    all_records.extend(index_records)

    stats = {
        "source_file": source_file,
        "record_count": len(all_records),
        "chunk_type_counts": dict(Counter(r["chunk_type"] for r in all_records)),
        "layouts": layouts,
        "code_map_count": len(code_map),
        "parser_version": "excel-rag-jsonl-parser-v1",
    }
    return all_records, stats


def write_jsonl(records: List[Dict[str, Any]], path: Path) -> None:
    with path.open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")


def write_markdown_sample(records: List[Dict[str, Any]], stats: Dict[str, Any], path: Path, sample_size: int = 12) -> None:
    lines = []
    lines.append("# Excel RAG JSONL Parser Output Summary\n")
    lines.append("## Stats\n")
    lines.append("```json")
    lines.append(json.dumps(stats, ensure_ascii=False, indent=2))
    lines.append("```\n")
    lines.append("## Sample Records\n")
    for record in records[:sample_size]:
        lines.append(f"### {record.get('chunk_type')} / {record.get('sheet')} / {record.get('excel_row', '')}\n")
        lines.append("```json")
        lines.append(json.dumps(record, ensure_ascii=False, indent=2))
        lines.append("```\n")
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse complex Excel workbook into RAG-ready JSONL chunks.")
    parser.add_argument("input", type=Path, help="Input .xlsx path")
    parser.add_argument("--output", type=Path, default=Path("rag_chunks.jsonl"), help="Output JSONL path")
    parser.add_argument("--stats", type=Path, default=Path("rag_chunks_stats.json"), help="Output stats JSON path")
    parser.add_argument("--sample-md", type=Path, default=Path("rag_chunks_sample.md"), help="Output markdown sample path")
    args = parser.parse_args()

    records, stats = parse_workbook(args.input)
    write_jsonl(records, args.output)
    args.stats.write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")
    write_markdown_sample(records, stats, args.sample_md)

    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

