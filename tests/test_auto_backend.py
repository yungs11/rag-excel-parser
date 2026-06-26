"""AutoBackend 2-tier 라우팅 (키워드 선필터 + delegation_rule 자기검증)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from excel_parser_rag.backends import get_backend
from excel_parser_rag.backends.auto_backend import detect_delegation_keyword, _should_try_openpyxl
from excel_parser_rag.config import ParserConfig


def _delegation_xlsx(path: Path) -> None:
    # Task 2 통합 테스트와 동일한 구조(헤더 style gate + 계층 행)로 작성해야
    # openpyxl backend 가 실제로 delegation matrix 로 분류해 delegation_rule 청크를 낸다.
    wb = Workbook(); ws = wb.active; ws.title = "위임전결"
    ws["A1"] = "위임전결 기준표"; ws.merge_cells("A1:D1")
    ws["A3"], ws["B3"], ws["C3"], ws["D3"] = "전결사항", "팀장", "본부장", "대표이사"
    for col in ("A", "B", "C", "D"):
        cell = ws[f"{col}3"]
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    data = [
        ("1. 업무전반", "", "", ""),
        ("라. 직원의 출장", "", "", ""),
        ("(1) 시외 출장", "", "", ""),
        ("(가) 팀장", "○", "", ""),
        ("(나) 팀원", "○", "", ""),
        ("마. 보고", "○", "", ""),
    ]
    r = 4
    for label, a, b, c in data:
        ws[f"A{r}"], ws[f"B{r}"], ws[f"C{r}"], ws[f"D{r}"] = label, a, b, c; r += 1
    wb.save(path)


def _flat_xlsx(path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "자산목록"
    ws["A1"], ws["B1"], ws["C1"] = "자산명", "담당자", "상태"
    ws["A2"], ws["B2"], ws["C2"] = "서버1", "홍길동", "사용"
    ws["A3"], ws["B3"], ws["C3"] = "서버2", "김철수", "미사용"
    wb.save(path)


def test_keyword_detect_true_for_delegation(tmp_path):
    p = tmp_path / "d.xlsx"; _delegation_xlsx(p)
    assert detect_delegation_keyword(p) is True


def test_keyword_detect_false_for_flat(tmp_path):
    p = tmp_path / "f.xlsx"; _flat_xlsx(p)
    assert detect_delegation_keyword(p) is False


def test_xls_suffix_never_routes_openpyxl(tmp_path):
    """`.xls` 는 (파일 미존재여도) openpyxl 라우팅 대상이 아님 — 단락 평가로 detect 미호출."""
    assert _should_try_openpyxl(tmp_path / "legacy.xls") is False
    assert _should_try_openpyxl(tmp_path / "legacy.XLS") is False


def test_auto_routes_delegation_to_openpyxl_merged(tmp_path):
    p = tmp_path / "d.xlsx"; _delegation_xlsx(p)
    cfg = ParserConfig(); cfg.backend = "auto"
    chunks, stats = get_backend("auto").parse(p, cfg)
    assert any(c["chunk_type"] == "delegation_rule" for c in chunks)
    assert stats.get("routed_backend") == "openpyxl"


def test_auto_routes_flat_to_kordoc(tmp_path):
    p = tmp_path / "f.xlsx"; _flat_xlsx(p)
    cfg = ParserConfig(); cfg.backend = "auto"
    cfg.kordoc_bin = "kordoc"; cfg.kordoc_md_out = "/tmp/kordoc_md"
    chunks, stats = get_backend("auto").parse(p, cfg)
    assert stats.get("routed_backend") == "kordoc"
    assert not any(c["chunk_type"] == "delegation_rule" for c in chunks)
