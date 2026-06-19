"""SoT §28.1 의 12종 테스트 fixture xlsx 빌더.

각 빌더는 작고 결정적인 한국어 데이터로 fixture 를 생성하고 저장 경로(Path)를 반환한다.
테스트가 기대값을 공유할 수 있도록 주요 상수를 모듈 레벨에 노출한다.

스타일 관례 (실제 위임전결기준표 분석 결과 반영):
- 헤더 행: bold + 가운데 정렬 + 연회색 solid fill(FFF3F3F3) + 풀테두리
- 대분류(섹션) 행: bold + 남색 solid fill(FF002060) + 흰 글씨
- 본문: 무스타일 또는 좌/우 테두리, 잔재(junk)는 무테두리
"""

from __future__ import annotations

from pathlib import Path
from typing import Callable, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_THIN = Side(style="thin")
FULL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
LR_BORDER = Border(left=_THIN, right=_THIN)
HEADER_FILL = PatternFill(fill_type="solid", start_color="FFF3F3F3", end_color="FFF3F3F3")
SECTION_FILL = PatternFill(fill_type="solid", start_color="FF002060", end_color="FF002060")
CENTER = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# 공유 기대값 상수 (테스트에서 import)
# ---------------------------------------------------------------------------

FLAT_01_SHEET = "실적"
FLAT_01_HEADERS = ["부서", "매출", "비용", "영업이익"]
FLAT_01_ROWS = [
    ("AI팀", 100000000, 70000000, 30000000),
    ("데이터팀", 80000000, 50000000, 30000000),
    ("플랫폼팀", 60000000, 45000000, 15000000),
    ("영업팀", 90000000, 72000000, 18000000),
]

MULTI_02_SHEET = "연도별실적"
MULTI_02_ROWS = [
    ("AI팀", 120, 30, 150, 45),
    ("데이터팀", 90, 20, 100, 25),
    ("플랫폼팀", 70, 15, 80, 18),
]

HIER_03_SHEET = "기준"

MATRIX_04_SHEET = "권한"
MATRIX_04_HEADERS = ["업무", "팀장", "본부장", "대표이사"]
# (행 라벨, {열 헤더: marker}) — marker 총 5개
MATRIX_04_BODY = [
    ("계정 생성", {"팀장": "○"}),
    ("계정 삭제", {"본부장": "○"}),
    ("시스템 재기동", {"대표이사": "●"}),
    ("예산 편성", {"본부장": "△"}),
    ("보안 점검", {"팀장": "○"}),
]
MATRIX_04_MARKER_COUNT = 5

HM_05_SHEET = "위임전결"

FORM_06_SHEET = "신청서"
FORM_06_TITLE = "출장 신청서"
FORM_06_FIELDS = {
    "신청자": "홍길동",
    "부서": "AI팀",
    "신청일자": "2026-04-17",
    "출장지": "부산광역시",
    "사유": "고객사 미팅 참석 및 기술 지원",
}

CODE_07_SHEET = "약어"
CODE_07_MAPPINGS = {
    "기": "기술관리실",
    "기획": "종합기획실",
    "인총": "인사총무실",
    "준감": "준법감시인",
    "HR": "인사팀",
}

MULTI_REGION_08_SHEET = "복합"
# (min_row, max_row) — 두 표의 행 구간
MULTI_REGION_08_SPANS = [(1, 4), (8, 11)]

NOTES_09_SHEET = "지급기준"
NOTES_09_NOTE_TEXT = "※ 해외출장의 경우 별도 기준을 따른다."
NOTES_09_NOTE_ROW = 5
NOTES_09_TOTAL_ROW = 4

HIDDEN_10_SHEET = "본문"
HIDDEN_10_HIDDEN_SHEET = "숨김시트"
HIDDEN_10_HIDDEN_ROW = 3   # "데이터팀" 행
HIDDEN_10_HIDDEN_COL = 3   # C열 "내부코드"

FORMULA_11_SHEET = "수식"

MESSY_12_SHEET = "전결기준"
MESSY_12_NOTE_TEXT = "※ 경미한 경우 전결 두 단계 하향"


# ---------------------------------------------------------------------------
# 스타일 헬퍼
# ---------------------------------------------------------------------------

def _style_header_range(ws, range_str: str) -> None:
    """병합 멤버(MergedCell) 포함 범위 전체에 헤더 스타일 적용 (병합 후 호출)."""
    for row in ws[range_str]:
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = CENTER
            cell.fill = HEADER_FILL
            cell.border = FULL_BORDER


def _style_section_cell(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFFFF")
    cell.fill = SECTION_FILL


def _center(cell) -> None:
    cell.alignment = CENTER


def _save(wb: Workbook, path) -> Path:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# 01 — 단순 flat table (부서별 실적표)
# ---------------------------------------------------------------------------

def build_01_flat_table(path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = FLAT_01_SHEET
    for j, header in enumerate(FLAT_01_HEADERS, start=1):
        ws.cell(row=1, column=j, value=header)
    _style_header_range(ws, "A1:D1")
    for i, row in enumerate(FLAT_01_ROWS, start=2):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    return _save(wb, path)


# ---------------------------------------------------------------------------
# 02 — 다단 헤더 표 (연도별 실적, 가로 병합 헤더)
# ---------------------------------------------------------------------------

def build_02_multi_header_table(path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = MULTI_02_SHEET
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:C1")
    ws.merge_cells("D1:E1")
    ws["A1"] = "구분"
    ws["B1"] = "2025"
    ws["D1"] = "2026"
    ws["B2"] = "매출"
    ws["C2"] = "영업이익"
    ws["D2"] = "매출"
    ws["E2"] = "영업이익"
    _style_header_range(ws, "A1:E2")
    for i, row in enumerate(MULTI_02_ROWS, start=3):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    return _save(wb, path)


# ---------------------------------------------------------------------------
# 03 — 병합 계층 표 (복리후생 기준, 세로 병합 + 번호 패턴)
# ---------------------------------------------------------------------------

def build_03_merged_hierarchy_table(path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = HIER_03_SHEET
    for j, header in enumerate(["구분", "항목", "세부항목", "기준"], start=1):
        ws.cell(row=1, column=j, value=header)
    _style_header_range(ws, "A1:D1")

    ws.merge_cells("A2:A4")
    ws["A2"] = "1. 임직원 복리후생"
    ws["A2"].font = Font(bold=True)
    ws.merge_cells("B2:B3")
    ws["B2"] = "가. 경조금"
    ws["C2"] = "(1) 결혼"
    ws["D2"] = "100만원"
    ws["C3"] = "(2) 출산"
    ws["D3"] = "50만원"
    ws["B4"] = "나. 학자금"
    ws["C4"] = "(1) 대학교"
    ws["D4"] = "300만원"

    ws.merge_cells("A5:A6")
    ws["A5"] = "2. 근태관리"
    ws["A5"].font = Font(bold=True)
    ws.merge_cells("B5:B6")
    ws["B5"] = "가. 연차"
    ws["C5"] = "(1) 신입"
    ws["D5"] = "11일"
    ws["C6"] = "(2) 경력"
    ws["D6"] = "15일"
    return _save(wb, path)


# ---------------------------------------------------------------------------
# 04 — matrix table (권한 매트릭스, marker 5개)
# ---------------------------------------------------------------------------

def build_04_matrix_table(path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = MATRIX_04_SHEET
    for j, header in enumerate(MATRIX_04_HEADERS, start=1):
        ws.cell(row=1, column=j, value=header)
    _style_header_range(ws, "A1:D1")
    col_of = {name: idx for idx, name in enumerate(MATRIX_04_HEADERS, start=1)}
    for i, (label, markers) in enumerate(MATRIX_04_BODY, start=2):
        ws.cell(row=i, column=1, value=label)
        for col_name, marker in markers.items():
            cell = ws.cell(row=i, column=col_of[col_name], value=marker)
            _center(cell)
    return _save(wb, path)


# ---------------------------------------------------------------------------
# 05 — hierarchical matrix (미니 위임전결표: 계층 + marker + 합의)
# ---------------------------------------------------------------------------

def build_05_hierarchical_matrix(path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = HM_05_SHEET
    ws.merge_cells("A1:B1")
    ws["A1"] = "전 결 사 항"
    ws["C1"] = "대표이사"
    ws["D1"] = "본부장"
    ws["E1"] = "팀장"
    ws["F1"] = "합의"
    _style_header_range(ws, "A1:F1")

    ws["A2"] = "1. 업무전반"
    _style_section_cell(ws["A2"])
    ws["A3"] = "가. 직원의 출장"
    ws["B4"] = "(1) 시외 출장"
    _center(ws.cell(row=4, column=4, value="○"))   # 본부장
    ws["B5"] = "(2) 시내 출장"
    _center(ws.cell(row=5, column=5, value="○"))   # 팀장
    ws["F5"] = "기획"

    ws["A6"] = "2. 인사"
    _style_section_cell(ws["A6"])
    ws["A7"] = "가. 채용 품의"
    _center(ws.cell(row=7, column=3, value="○"))   # 대표이사
    ws["F7"] = "인총"
    return _save(wb, path)


# ---------------------------------------------------------------------------
# 06 — form 문서 (출장 신청서, key-value)
# ---------------------------------------------------------------------------

def build_06_form_document(path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = FORM_06_SHEET
    ws.merge_cells("A1:D1")
    ws["A1"] = FORM_06_TITLE
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = CENTER

    ws["A2"] = "신청자"
    ws["B2"] = FORM_06_FIELDS["신청자"]
    ws["C2"] = "부서"
    ws["D2"] = FORM_06_FIELDS["부서"]
    ws["A3"] = "신청일자"
    ws["B3"] = FORM_06_FIELDS["신청일자"]
    ws["C3"] = "출장지"
    ws["D3"] = FORM_06_FIELDS["출장지"]
    ws["A4"] = "사유"
    ws.merge_cells("B4:D4")
    ws["B4"] = FORM_06_FIELDS["사유"]
    for coord in ("A2", "C2", "A3", "C3", "A4"):
        cell = ws[coord]
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = FULL_BORDER
    return _save(wb, path)


# ---------------------------------------------------------------------------
# 07 — code mapping (약어 → 정식명칭)
# ---------------------------------------------------------------------------

def build_07_code_mapping(path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = CODE_07_SHEET
    ws["A1"] = "약어"
    ws["B1"] = "정식명칭"
    _style_header_range(ws, "A1:B1")
    for i, (code, meaning) in enumerate(CODE_07_MAPPINGS.items(), start=2):
        a = ws.cell(row=i, column=1, value=code)
        b = ws.cell(row=i, column=2, value=meaning)
        a.border = FULL_BORDER
        b.border = FULL_BORDER
    return _save(wb, path)


# ---------------------------------------------------------------------------
# 08 — 한 시트에 표 2개 (빈 행 3줄로 분리)
# ---------------------------------------------------------------------------

def build_08_multiple_regions_one_sheet(path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = MULTI_REGION_08_SHEET
    # 표 1: rows 1-4
    ws["A1"] = "부서"
    ws["B1"] = "인원"
    _style_header_range(ws, "A1:B1")
    for i, (dept, count) in enumerate(
        [("AI팀", 12), ("데이터팀", 8), ("영업팀", 15)], start=2
    ):
        ws.cell(row=i, column=1, value=dept)
        ws.cell(row=i, column=2, value=count)
    # rows 5-7 빈 행
    # 표 2: rows 8-11
    ws["A8"] = "항목"
    ws["B8"] = "금액"
    _style_header_range(ws, "A8:B8")
    for i, (item, amount) in enumerate(
        [("출장비", 500000), ("회식비", 300000), ("교육비", 700000)], start=9
    ):
        ws.cell(row=i, column=1, value=item)
        ws.cell(row=i, column=2, value=amount)
    return _save(wb, path)


# ---------------------------------------------------------------------------
# 09 — note / total 행 포함 표
# ---------------------------------------------------------------------------

def build_09_notes_and_footers(path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = NOTES_09_SHEET
    ws["A1"] = "항목"
    ws["B1"] = "지급액"
    _style_header_range(ws, "A1:B1")
    ws["A2"] = "출장비(국내)"
    ws["B2"] = 50000
    ws["A3"] = "출장비(해외)"
    ws["B3"] = 150000
    ws["A4"] = "합계"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = 200000
    ws["A5"] = NOTES_09_NOTE_TEXT
    return _save(wb, path)


# ---------------------------------------------------------------------------
# 10 — 숨김 행/열 + 숨김 시트
# ---------------------------------------------------------------------------

def build_10_hidden_rows_cols(path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = HIDDEN_10_SHEET
    ws["A1"] = "부서"
    ws["B1"] = "매출"
    ws["C1"] = "내부코드"
    _style_header_range(ws, "A1:C1")
    for i, (dept, sales, code) in enumerate(
        [("AI팀", 100, "A01"), ("데이터팀", 80, "D02"), ("비공개팀", 50, "X09")], start=2
    ):
        ws.cell(row=i, column=1, value=dept)
        ws.cell(row=i, column=2, value=sales)
        ws.cell(row=i, column=3, value=code)
    ws.row_dimensions[HIDDEN_10_HIDDEN_ROW].hidden = True   # "데이터팀" 행
    ws.column_dimensions["C"].hidden = True                  # "내부코드" 열

    hidden_ws = wb.create_sheet(HIDDEN_10_HIDDEN_SHEET)
    hidden_ws["A1"] = "숨김 데이터"
    hidden_ws.sheet_state = "hidden"
    return _save(wb, path)


# ---------------------------------------------------------------------------
# 11 — 수식 값 (openpyxl 생성 파일은 캐시값이 없음에 유의)
# ---------------------------------------------------------------------------

def build_11_formula_values(path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = FORMULA_11_SHEET
    ws["A1"] = "항목"
    ws["B1"] = "금액"
    _style_header_range(ws, "A1:B1")
    ws["A2"] = "기본급"
    ws["B2"] = 3000000
    ws["A3"] = "수당"
    ws["B3"] = 500000
    ws["A4"] = "합계"
    ws["B4"] = "=B2+B3"   # data_only 로드 시 캐시값 없음 → None
    return _save(wb, path)


# ---------------------------------------------------------------------------
# 12 — messy real world (제목/빈 행/계층 매트릭스/note/잔재 혼합)
# ---------------------------------------------------------------------------

def build_12_messy_real_world(path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = MESSY_12_SHEET

    ws["A1"] = "<별표1>"
    ws.merge_cells("B2:F2")
    ws["B2"] = "위임전결 기준표(테스트)"
    ws["B2"].font = Font(bold=True, size=14)
    ws["B2"].alignment = CENTER
    # rows 3-4: 빈 행

    ws.merge_cells("B5:C5")
    ws["B5"] = "전 결 사 항"
    ws["D5"] = "대표이사"
    ws["E5"] = "본부장"
    ws["F5"] = "팀장"
    ws["G5"] = "합의"
    _style_header_range(ws, "B5:G5")

    ws["B6"] = "1. 업무전반"
    _style_section_cell(ws["B6"])
    ws["B7"] = "가. 직원의 출장"
    ws["C8"] = "(1) 시외 출장"
    _center(ws.cell(row=8, column=5, value="○"))   # 본부장
    ws["C9"] = "(2) 시내 출장"
    _center(ws.cell(row=9, column=6, value="○"))   # 팀장
    ws["G9"] = "기(준감)"
    ws["B10"] = MESSY_12_NOTE_TEXT
    ws["B11"] = "나. 휴가 승인"
    _center(ws.cell(row=11, column=6, value="○"))  # 팀장

    ws["B12"] = "2. 인사"
    _style_section_cell(ws["B12"])
    ws["B13"] = "가. 채용 품의"
    _center(ws.cell(row=13, column=4, value="○"))  # 대표이사
    ws["G13"] = "인총"

    # 본문 좌/우 테두리 (정상 본문 vs 잔재 구분 신호)
    for r in range(6, 14):
        for c in range(2, 8):
            ws.cell(row=r, column=c).border = LR_BORDER

    # rows 14-19 빈 행 후 잔재 (무테두리 고아 셀)
    ws["D20"] = "잔재 텍스트 조각"
    return _save(wb, path)


# ---------------------------------------------------------------------------
# 레지스트리
# ---------------------------------------------------------------------------

FIXTURE_BUILDERS: Dict[str, Callable[[Path], Path]] = {
    "01_flat_table.xlsx": build_01_flat_table,
    "02_multi_header_table.xlsx": build_02_multi_header_table,
    "03_merged_hierarchy_table.xlsx": build_03_merged_hierarchy_table,
    "04_matrix_table.xlsx": build_04_matrix_table,
    "05_hierarchical_matrix.xlsx": build_05_hierarchical_matrix,
    "06_form_document.xlsx": build_06_form_document,
    "07_code_mapping.xlsx": build_07_code_mapping,
    "08_multiple_regions_one_sheet.xlsx": build_08_multiple_regions_one_sheet,
    "09_notes_and_footers.xlsx": build_09_notes_and_footers,
    "10_hidden_rows_cols.xlsx": build_10_hidden_rows_cols,
    "11_formula_values.xlsx": build_11_formula_values,
    "12_messy_real_world.xlsx": build_12_messy_real_world,
}

FIXTURE_NAMES: List[str] = list(FIXTURE_BUILDERS.keys())
