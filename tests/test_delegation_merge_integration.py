"""DelegationRulePlugin + sibling_merger 통합 (합성 전결 xlsx)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from excel_parser_rag.backends import get_backend
from excel_parser_rag.config import ParserConfig


def _make_delegation_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "위임전결"
    ws["A1"] = "위임전결 기준표"
    ws.merge_cells("A1:D1")
    # 헤더: 전결사항 | 팀장 | 본부장 | 대표이사
    ws["A3"], ws["B3"], ws["C3"], ws["D3"] = "전결사항", "팀장", "본부장", "대표이사"
    # header_detector 의 style gate(bold/fill) 통과용 — 헤더 행 스타일 지정
    for col in ("A", "B", "C", "D"):
        cell = ws[f"{col}3"]
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    rows = [
        ("1. 업무전반", "", "", ""),            # 섹션 헤더 행
        ("라. 직원의 출장", "", "", ""),         # 부모
        ("(1) 시외 출장", "", "", ""),           # 하위부모
        ("(가) 팀장", "○", "", ""),              # 형제 leaf 1
        ("(나) 팀원", "○", "", ""),              # 형제 leaf 2
        ("마. 보고", "○", "", ""),               # 다른 항목
    ]
    r = 4
    for label, a, b, c in rows:
        ws[f"A{r}"], ws[f"B{r}"], ws[f"C{r}"], ws[f"D{r}"] = label, a, b, c
        r += 1
    wb.save(path)


def test_openpyxl_delegation_produces_merged_chunk(tmp_path):
    p = tmp_path / "deleg.xlsx"
    _make_delegation_xlsx(p)
    cfg = ParserConfig()
    cfg.backend = "openpyxl"
    cfg.delegation_merge_max_chars = 1100
    chunks, _stats = get_backend("openpyxl").parse(p, cfg)
    deleg = [c for c in chunks if c["chunk_type"] == "delegation_rule"]
    assert deleg, "delegation_rule 청크가 없음"
    merged = [c for c in deleg if c.get("metadata", {}).get("merged")]
    assert merged, f"병합 청크 없음. delegation_rule={len(deleg)}"
    # (가)팀장 + (나)팀원 이 한 병합 청크에 함께
    target = [c for c in merged if "(가) 팀장" in c["content_text"] and "(나) 팀원" in c["content_text"]]
    assert target, f"형제 병합 실패: {[c['content_text'][:60] for c in merged]}"
    assert len(target[0]["content_text"]) <= 1100


def test_merge_disabled_when_max_chars_zero(tmp_path):
    p = tmp_path / "deleg2.xlsx"
    _make_delegation_xlsx(p)
    cfg = ParserConfig()
    cfg.backend = "openpyxl"
    cfg.delegation_merge_max_chars = 0  # 비활성
    chunks, _ = get_backend("openpyxl").parse(p, cfg)
    merged = [c for c in chunks if c.get("metadata", {}).get("merged")]
    assert not merged, "max_chars=0 인데 병합이 일어남"
