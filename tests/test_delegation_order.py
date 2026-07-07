"""delegation 플러그인 출력이 엑셀 위→아래 행 순서로 정렬되는지 검증.

버그: base_chunks(note 등)를 먼저, delegation_rows 를 뒤에 append → [note 블록][delegation 블록]
2블록으로 뒤섞임. 수정: parse() 반환을 (타입클래스, start_row) 안정정렬.
"""
from __future__ import annotations

from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from excel_parser_rag.backends import get_backend
from excel_parser_rag.config import ParserConfig


def _make_fixture(path) -> None:
    """전결표 + delegation 행들보다 뒤(높은 행)에 ※주석 행. 헤더행 스타일링(전결 인식 조건)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "위임전결"
    ws["A1"] = "위임전결 기준표"
    ws.merge_cells("A1:D1")
    ws["A3"], ws["B3"], ws["C3"], ws["D3"] = "전결사항", "팀장", "본부장", "대표이사"
    fill = PatternFill("solid", fgColor="D9D9D9")
    for col in ("A", "B", "C", "D"):
        cell = ws[f"{col}3"]
        cell.font = Font(bold=True)
        cell.fill = fill
    rows = [
        (4, "1. 업무전반", "", "", ""),
        (5, "가. 처리원칙", "○", "", ""),
        (6, "나. 부서업무", "○", "", ""),
        (7, "다. 소속팀원", "○", "", ""),
        (10, "※ 본 기준은 예시임", "", "", ""),  # 주석 — delegation 행들보다 뒤(행10)
    ]
    for r, a, b, c, d in rows:
        ws[f"A{r}"], ws[f"B{r}"], ws[f"C{r}"], ws[f"D{r}"] = a, b, c, d
    wb.save(path)


def _parse(path):
    cfg = ParserConfig()
    cfg.backend = "openpyxl"
    return get_backend("openpyxl").parse(path, cfg)[0]


def _row(c):
    return (c.get("source") or {}).get("start_row")


def test_note_interleaves_by_row_not_front_block(tmp_path):
    """수정 전(2블록)엔 모든 note 가 모든 delegation 앞 → 실패. 수정 후 note 가 앞선 행 delegation 뒤 → 통과."""
    p = tmp_path / "deleg.xlsx"
    _make_fixture(p)
    chunks = _parse(p)
    notes = [(i, c) for i, c in enumerate(chunks) if c["chunk_type"] == "note"]
    dels = [(i, c) for i, c in enumerate(chunks) if c["chunk_type"] == "delegation_rule"]
    assert notes and dels, f"note/delegation 미생성: {dict(Counter(c['chunk_type'] for c in chunks))}"
    ni, nc = notes[0]
    # 어떤 delegation 의 index 가 note 보다 앞이고 그 start_row 도 note 보다 작다 = 행순 인터리브.
    assert any(di < ni and _row(dc) < _row(nc) for di, dc in dels), (
        "note 가 여전히 앞선 행 delegation 보다 앞(2블록) — 행순 정렬 안 됨"
    )


def test_body_chunks_row_ordered(tmp_path):
    """body 청크(요약 제외)의 start_row 가 비감소(역행 0). 0-fallback 위양성 차단 위해 int 보유 선단언."""
    p = tmp_path / "deleg2.xlsx"
    _make_fixture(p)
    chunks = _parse(p)
    body = [c for c in chunks if c["chunk_type"] not in ("table_summary", "section_summary")]
    assert body, "body 청크 없음"
    for c in body:
        assert isinstance(_row(c), int), f"start_row 없음: {c['chunk_type']}"
    rows = [_row(c) for c in body]
    assert rows == sorted(rows), f"body 행 역행: {rows}"


def test_summary_class_placement(tmp_path):
    """table_summary=맨 앞, section_summary=맨 뒤(그 뒤에 body 없음)."""
    p = tmp_path / "deleg3.xlsx"
    _make_fixture(p)
    chunks = _parse(p)
    types = [c["chunk_type"] for c in chunks]
    if "table_summary" in types:
        assert types.index("table_summary") == 0
    if "section_summary" in types:
        first_sec = types.index("section_summary")
        assert all(t == "section_summary" for t in types[first_sec:]), (
            f"section_summary 뒤에 body 청크: {types[first_sec:]}"
        )
