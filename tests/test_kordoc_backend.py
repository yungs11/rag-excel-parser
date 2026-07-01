"""kordoc 백엔드 통합 테스트 (합성 xlsx + kordoc md 페어)."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_parser_rag.backends import get_backend
from excel_parser_rag.backends.base import BackendError
from excel_parser_rag.config import ParserConfig
from excel_parser_rag.chunking.chunk_schema import validate_chunk_schema


def _make_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws["A1"] = "테스트표"
    ws.merge_cells("A1:C1")
    ws.append([])  # 빈 행은 자동으로 안 들어가므로 직접 좌표 기입
    ws["A3"], ws["B3"], ws["C3"] = "항목", "담당", "비고"
    ws["A4"], ws["B4"], ws["C4"] = "작업A", "홍길동", "메모1"
    ws["A5"], ws["B5"], ws["C5"] = "작업B", "김철수", "메모2"
    # marker 매트릭스 행
    ws["A6"], ws["D6"] = "승인건", "○"
    ws["D3"] = "팀장"
    wb.save(path)


MD = """## S1

<table>
<tr><th colspan="4">테스트표</th></tr>
<tr><td>항목</td><td>담당</td><td>비고</td><td>팀장</td></tr>
<tr><td>작업A</td><td>홍길동</td><td>메모1</td><td></td></tr>
<tr><td>작업B</td><td>김철수</td><td>메모2</td><td></td></tr>
<tr><td>승인건</td><td></td><td></td><td>○</td></tr>
</table>
"""


@pytest.fixture()
def doc(tmp_path):
    xlsx = tmp_path / "테스트표.xlsx"
    md = tmp_path / "테스트표.md"
    _make_xlsx(xlsx)
    md.write_text(MD, encoding="utf-8")
    return xlsx, md


def test_kordoc_backend_basic(doc):
    xlsx, md = doc
    cfg = ParserConfig(backend="kordoc", kordoc_md_path=str(md))
    chunks, stats = get_backend("kordoc").parse(xlsx, cfg)

    assert stats["backend"] == "kordoc"
    assert len(chunks) >= 3
    # 모든 청크 좌표 + 스키마
    assert all(c["range"] for c in chunks)
    for c in chunks:
        assert not validate_chunk_schema(c), c.get("id")
    # 시트제목이 섹션으로 새지 않고 title 로 분리
    assert all(c["title"] == "테스트표" for c in chunks)


def test_kordoc_backend_rows_and_coords(doc):
    xlsx, md = doc
    cfg = ParserConfig(backend="kordoc", kordoc_md_path=str(md))
    chunks, _ = get_backend("kordoc").parse(xlsx, cfg)
    rows = {c["range"]: c for c in chunks if c["chunk_type"] == "table_row"}
    # 작업A 행이 원본 A4:* 로 좌표 복원. 빈 헤더 컬럼(팀장)도 빈값으로 포함된다.
    a = next(c for c in chunks if c["fields"].get("항목") == "작업A")
    assert a["fields"] == {"항목": "작업A", "담당": "홍길동", "비고": "메모1", "팀장": ""}
    assert a["source"]["start_row"] == 4


# ─── 결함 회귀: (1) 시트명 중복 (2) 값 짤림 (3) 빈칸/중복헤더 ────────────
_WIDE_VALUE = "가" * 1000  # content 600 / embedding 900 상한을 초과시키는 긴 값

MD2 = """## 자산

<table>
<tr><th colspan="5">자산</th></tr>
<tr><td>ID</td><td>담당자</td><td>메모</td><td>담당자</td><td>도입일</td></tr>
<tr><td>SV-1</td><td>홍길동</td><td>__WIDE__</td><td></td><td>2019.07</td></tr>
</table>
""".replace("__WIDE__", _WIDE_VALUE)


@pytest.fixture()
def doc2(tmp_path):
    xlsx = tmp_path / "자산.xlsx"
    md = tmp_path / "자산.md"
    wb = Workbook()
    ws = wb.active
    ws.title = "자산"
    ws["A1"] = "자산"
    ws.merge_cells("A1:E1")
    ws["A3"], ws["B3"], ws["C3"], ws["D3"], ws["E3"] = "ID", "담당자", "메모", "담당자", "도입일"
    ws["A4"], ws["B4"], ws["C4"], ws["E4"] = "SV-1", "홍길동", _WIDE_VALUE, "2019.07"
    wb.save(xlsx)
    md.write_text(MD2, encoding="utf-8")
    return xlsx, md


def test_issue1_sheet_name_not_duplicated(doc2):
    """title == sheet 일 때 '자산의 자산 시트' 중복 금지 → '자산 시트'."""
    xlsx, md = doc2
    cfg = ParserConfig(backend="kordoc", kordoc_md_path=str(md))
    chunks, _ = get_backend("kordoc").parse(xlsx, cfg)
    row = next(c for c in chunks if c["fields"].get("ID") == "SV-1")
    assert "자산의 자산 시트" not in row["content_text"]
    assert row["content_text"].startswith("자산 시트")


def test_issue2_long_row_not_truncated(doc2):
    """넉넉한 상한(기본 content 3000 / embedding 4000)으로 긴 행이 안 짤린다."""
    xlsx, md = doc2
    cfg = ParserConfig(backend="kordoc", kordoc_md_path=str(md))
    chunks, _ = get_backend("kordoc").parse(xlsx, cfg)
    row = next(c for c in chunks if c["fields"].get("ID") == "SV-1")
    assert row["fields"]["메모"] == _WIDE_VALUE  # 값 자체는 온전
    assert _WIDE_VALUE in row["metadata"]["embedding_text"]  # 임베딩 텍스트에 전체 포함


def test_issue3_empty_and_duplicate_headers(doc2):
    """빈칸 헤더 포함 + 중복 라벨(담당자 2회)을 접미사로 구분."""
    xlsx, md = doc2
    cfg = ParserConfig(backend="kordoc", kordoc_md_path=str(md))
    chunks, _ = get_backend("kordoc").parse(xlsx, cfg)
    row = next(c for c in chunks if c["fields"].get("ID") == "SV-1")
    f = row["fields"]
    # 첫 담당자는 값, 두 번째 담당자(D열)는 빈칸이지만 접미사로 구분되어 포함
    assert f["담당자"] == "홍길동"
    assert f["담당자(D)"] == ""
    # 빈칸 도입일은 아니지만, 5개 헤더 전부 존재해야 함
    assert set(f) == {"ID", "담당자", "메모", "담당자(D)", "도입일"}


def test_kordoc_backend_compact_matrix(doc):
    xlsx, md = doc
    cfg = ParserConfig(backend="kordoc", kordoc_md_path=str(md))
    chunks, _ = get_backend("kordoc").parse(xlsx, cfg)
    mf = [c for c in chunks if c["chunk_type"] == "matrix_fact"]
    assert mf, "marker 행이 matrix_fact 로"
    m = mf[0]
    # compact: 마커를 "해당: <컬럼헤더>" 로 접음 (○ 노이즈 없음)
    assert m["fields"].get("해당") == "팀장"
    assert "○" not in m["fields"].get("해당", "")


def test_kordoc_backend_missing_md_errors(tmp_path):
    xlsx = tmp_path / "없는md.xlsx"
    _make_xlsx(xlsx)
    cfg = ParserConfig(backend="kordoc")  # md 미지정
    with pytest.raises(BackendError):
        get_backend("kordoc").parse(xlsx, cfg)
