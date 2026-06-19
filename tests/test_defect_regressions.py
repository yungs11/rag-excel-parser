"""검증 라운드에서 확인된 결함들의 회귀 테스트.

1. 'N-M.' 복합 번호 섹션(5-2/5-3)이 최상위로 복원되는지 (path 오염 방지)
2. Index 시트 CPO/CIAP/CISO 약어 매핑 추출
3. 스타일 없는 일반 flat table 이 code_mapping_table 로 오인되지 않는지
4. 숫자 0 이 marker('해당')로 둔갑하지 않는지
5. 제목 블록만 있는 시트가 통째로 누락되지 않는지
"""

from __future__ import annotations

import json

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font

from conftest import chunk_blob


# ---------------------------------------------------------------------------
# 헬퍼: 임시 워크북 생성 → parse
# ---------------------------------------------------------------------------

def _parse(tmp_path, name, build):
    from excel_parser_rag import ExcelRagParser

    wb = Workbook()
    build(wb)
    path = tmp_path / name
    wb.save(path)
    return ExcelRagParser().parse(path)


# ---------------------------------------------------------------------------
# 1. 'N-M.' 복합 번호 = 최상위 섹션 (hierarchy 단위)
# ---------------------------------------------------------------------------

class TestCompoundNumberedSectionLevel:
    def test_item_numbering_level_top(self):
        from excel_parser_rag.parsers.hierarchy_table import item_numbering_level

        assert item_numbering_level("5-2. ICT개발") == 0
        assert item_numbering_level("5-3. ICT 인프라") == 0
        assert item_numbering_level("10-1. 기타") == 0
        # 기존 패턴은 그대로
        assert item_numbering_level("5. ICT") == 0
        assert item_numbering_level("가. 항목") == 1
        assert item_numbering_level("(1) 항목") == 2
        assert item_numbering_level("일반 텍스트") is None

    def test_tracker_treats_compound_number_as_sibling(self):
        from excel_parser_rag.parsers.hierarchy_table import HierarchyTracker

        tracker = HierarchyTracker([2, 3])
        tracker.push("5-1. ICT 기획", 2)
        tracker.push("바. ICT정책 제개정", 2)
        tracker.push("(가) 제·개정", 3)
        path = tracker.push("5-2. ICT개발", 2)
        assert path == ["5-2. ICT개발"], f"5-2 는 5-1 의 형제(최상위)여야 함: {path}"


# ---------------------------------------------------------------------------
# 2~5. 실제 위임전결기준표 회귀 (파일 있을 때만)
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def real_chunks(real_excel_path):
    from excel_parser_rag import ExcelRagParser

    return ExcelRagParser().parse(real_excel_path)


class TestRealWorkbookRegressions:
    def test_sections_5_2_and_5_3_are_roots(self, real_chunks):
        """rows 259-305 의 chunk 는 path[0] 가 5-2/5-3 이어야 한다 (5-1 오염 금지)."""
        for ch in real_chunks:
            src = ch.get("source", {})
            if src.get("sheet") != "1.위임전결기준" or ch["chunk_type"] == "note":
                continue
            row = src.get("start_row")
            if not isinstance(row, int) or not ch.get("path"):
                continue
            if 259 <= row <= 273:
                assert ch["path"][0] == "5-2. ICT개발", (row, ch["chunk_type"], ch["path"][:3])
            elif 274 <= row <= 305:
                assert ch["path"][0] == "5-3. ICT 인프라", (row, ch["chunk_type"], ch["path"][:3])

    def test_section_summaries_for_5_2_and_5_3(self, real_chunks):
        names = {
            ch["fields"].get("섹션")
            for ch in real_chunks
            if ch["chunk_type"] == "section_summary"
            and ch.get("source", {}).get("sheet") == "1.위임전결기준"
        }
        assert "5-1. ICT 기획" in names
        assert "5-2. ICT개발" in names
        assert "5-3. ICT 인프라" in names

    def test_index_role_mappings_extracted(self, real_chunks):
        """Index 시트 E/F 열의 CPO/CIAP/CISO 역할 매핑이 살아남아야 한다."""
        mappings = [ch for ch in real_chunks if ch["chunk_type"] == "code_mapping"]
        blob = " ".join(chunk_blob(ch) for ch in mappings)
        for role in ("CEO", "CPO", "CIAP", "CISO"):
            assert role in blob, f"code_mapping 에 '{role}' 매핑 누락"
        pairs = {(ch["fields"]["약어"], ch["fields"]["의미"]) for ch in mappings}
        assert ("CISO(정보보호최고책임자)", "디지털혁신팀장") in pairs
        assert ("CPO(개인정보보호책임자)", "준법감시인") in pairs

    def test_ambiguous_o_markers_still_flagged(self, real_chunks):
        """3.정보보호 의 영문자 'O' marker 는 여전히 감점/경고 대상이어야 한다."""
        amb = [
            ch for ch in real_chunks
            if ch.get("metadata", {}).get("ambiguous_marker")
        ]
        assert amb, "ambiguous marker 경고가 모두 사라짐 (검증 약화)"
        assert all(ch["source"]["sheet"] == "3.정보보호" for ch in amb)


# ---------------------------------------------------------------------------
# 3. 일반 flat table 오분류 (code_mapping_table 사실 조작) 방지
# ---------------------------------------------------------------------------

class TestGenericFlatTableNotCodeMapping:
    def test_unstyled_roster_table(self, tmp_path):
        """스타일 없는 4x3 명단 표 — 코드표로 오인 금지 + 3번째 열 보존."""
        def build(wb):
            ws = wb.active
            ws.title = "명단"
            rows = [
                ("이름", "부서", "직급"),
                ("김철수", "재무팀", "과장"),
                ("이영희", "인사팀", "대리"),
                ("박민수", "총무팀", "차장"),
            ]
            for i, row in enumerate(rows, start=1):
                for j, v in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=v)

        chunks = _parse(tmp_path, "roster.xlsx", build)
        assert chunks, "chunk 0개 — 시트 누락"
        assert not any(c["chunk_type"] == "code_mapping" for c in chunks), (
            "일반 명단 표가 code_mapping 으로 사실을 지어냄"
        )
        blob = " ".join(chunk_blob(c) for c in chunks)
        assert "의미는" not in blob
        for v in ("과장", "대리", "차장"):  # 3번째 열 무손실
            assert v in blob, f"직급 열 값 '{v}' 누락 (silent column loss)"

    def test_bold_header_one_data_row(self, tmp_path):
        """헤더 1행 + 본문 1행 표 — flat_table 로 처리되어야 한다."""
        def build(wb):
            ws = wb.active
            ws.title = "실적"
            for j, h in enumerate(("부서", "매출", "비용"), start=1):
                cell = ws.cell(row=1, column=j, value=h)
                cell.font = Font(bold=True)
            for j, v in enumerate(("AI팀", 100, 70), start=1):
                ws.cell(row=2, column=j, value=v)

        chunks = _parse(tmp_path, "bold2row.xlsx", build)
        assert chunks
        assert not any(c["chunk_type"] == "code_mapping" for c in chunks)
        rows = [c for c in chunks if c["chunk_type"] == "table_row"]
        assert rows and rows[0]["region_type"] == "flat_table"
        blob = chunk_blob(rows[0])
        for v in ("AI팀", "100", "70"):
            assert v in blob


# ---------------------------------------------------------------------------
# 4. 숫자 0 은 marker 가 아니다 (SoT §8.1)
# ---------------------------------------------------------------------------

class TestZeroIsNotMarker:
    def test_numeric_zero_table(self, tmp_path):
        def build(wb):
            ws = wb.active
            ws.title = "재고"
            data = [
                ("품목", "1월", "2월", "3월"),
                ("1. 노트북", 0, 5, 3),
                ("2. 모니터", 2, 0, 0),
                ("3. 키보드", 0, 0, 7),
                ("4. 마우스", 1, 2, 0),
            ]
            for i, row in enumerate(data, start=1):
                for j, v in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=v)

        chunks = _parse(tmp_path, "stock.xlsx", build)
        assert chunks
        blob = " ".join(chunk_blob(c) for c in chunks)
        assert "해당" not in blob, "숫자 0 이 '해당' marker 로 둔갑"
        assert "applicable" not in blob
        # marker_ratio 가 0 이므로 매트릭스 계열로 분류되면 안 됨
        assert not any("matrix" in c["region_type"] for c in chunks)

    def test_markerutil_guards(self):
        from excel_parser_rag.canvas.cell_node import CellNode
        from excel_parser_rag.markerutil import (
            is_ambiguous_marker_cell,
            is_marker_cell,
            normalize_marker_cell,
        )

        zero = CellNode(sheet="s", row=1, col=1, raw_value=0, data_type="int")
        assert not is_marker_cell(zero, "0")
        assert normalize_marker_cell(zero, "0") is None
        assert not is_ambiguous_marker_cell(zero, "0")
        # 영문자 O 는 여전히 (모호) marker
        oh = CellNode(sheet="s", row=1, col=2, raw_value="O", data_type="str")
        assert is_marker_cell(oh, "O")
        assert normalize_marker_cell(oh, "O") == "applicable"
        assert is_ambiguous_marker_cell(oh, "O")
        circle = CellNode(sheet="s", row=1, col=3, raw_value="○", data_type="str")
        assert is_marker_cell(circle, "○")
        assert normalize_marker_cell(circle, "○") == "applicable"


# ---------------------------------------------------------------------------
# 5. 제목 블록만 있는 시트도 chunk 를 남긴다
# ---------------------------------------------------------------------------

class TestTitleOnlySheetNotDropped:
    def test_title_only_sheets_emit_chunks(self, tmp_path):
        def build(wb):
            ws = wb.active
            ws.title = "OneRow"
            ws["A1"] = "분기 보고"
            ws["B1"] = "2026년 1분기"
            ws2 = wb.create_sheet("Para")
            ws2["A1"] = (
                "이 문서는 2026년 1분기 운영 현황을 요약한 것으로, "
                "자세한 내용은 본문 보고서를 참조해야 한다."
            )

        chunks = _parse(tmp_path, "titleonly.xlsx", build)
        sheets = {c["source"]["sheet"] for c in chunks}
        assert "OneRow" in sheets, "OneRow 시트가 통째로 누락됨"
        assert "Para" in sheets, "Para 시트가 통째로 누락됨"
        blob = " ".join(chunk_blob(c) for c in chunks)
        assert "분기 보고" in blob
        assert "운영 현황" in blob

    def test_title_above_table_still_skipped(self, tmp_path):
        """표 위에 붙은 제목 영역은 여전히 chunk 미생성 (기존 동작 유지)."""
        def build(wb):
            ws = wb.active
            ws.title = "본문"
            ws.merge_cells("A1:C1")
            cell = ws["A1"]
            cell.value = "부서별 인원 현황"
            cell.font = Font(bold=True, size=14)
            for j, h in enumerate(("부서", "인원", "비고"), start=1):
                c = ws.cell(row=3, column=j, value=h)
                c.font = Font(bold=True)
            for i, row in enumerate([("AI팀", 12, "-"), ("데이터팀", 8, "-")], start=4):
                for j, v in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=v)

        chunks = _parse(tmp_path, "title_table.xlsx", build)
        assert chunks
        # 제목 텍스트만 담은 별도 영역 chunk 가 표 chunk 와 중복 생성되지 않아야 함
        title_only = [
            c for c in chunks
            if c["chunk_type"] == "table_row" and c["fields"].get("A") == "부서별 인원 현황"
        ]
        assert not title_only
