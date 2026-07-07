"""게이트 검증 요약 — 추출 실패(ref/header_leak/empty_header) + 나란히2표.

설계: docs/superpowers/specs/2026-06-29-excel-gate-postparse-design.md
백엔드(openpyxl/kordoc) 무관하게 동작: 원시 셀(openpyxl)로 구조/참조,
실제 파싱 chunks 로 헤더누수를 판정한다.
"""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.utils import get_column_letter

from ..config import ParserConfig
from ..pipeline import build_canvases, detect_and_classify
from ..parsers.hierarchy_table import item_numbering_level
from ..parsers.flat_table import cell_text, body_rows_of

ERROR_RE = re.compile(r"#(REF|VALUE|DIV/0|N/A|NAME\?|NULL|NUM)!?")
# 인덱스열 라벨(나란히 놓인 독립 표 각각의 행번호 열) — 중복 시 side_by_side 강신호
_INDEX_RE = re.compile(r"^(순번|연번|번호|no\.?|#|seq|id)$", re.IGNORECASE)


def _numbering_restart(col_min_ordered):
    """계층열 왼쪽→오른쪽 순서의 [(col, min_level), ...] 에서, 깊은 열의 최소 번호레벨이
    얕은 열 이하(<=)로 되돌아가는(번호 재시작) 첫 쌍 (cs, cd) 을 반환. 없으면 None.
    각 열의 최소레벨이 strictly 증가해야 정상 계층(2-1 위임전결형); 되돌아가면 상하위 모호."""
    for i in range(1, len(col_min_ordered)):
        (cs, ms), (cd, md) = col_min_ordered[i - 1], col_min_ordered[i]
        if md <= ms:
            return (cs, cd)
    return None


def _detect_side_by_side(labels: Dict[int, str]):
    """나란히 놓인 두 표 판정 → 관련 컬럼 set.
    (A) 인덱스열(순번/No 등) 라벨이 2회 이상 등장, 또는
    (B) ≥2개 라벨로 된 연속 블록이 헤더행에서 통째로 반복.
    단순 단일 비인덱스 라벨 1회 중복(한 표의 동명 컬럼)이나 매트릭스(사람별 열)는 제외.
    """
    ordered = sorted(labels.items())  # [(col, label), ...] 열 순서
    cols = [c for c, _ in ordered]
    seq = [lab for _, lab in ordered]
    sbs_cols: set = set()
    counts = Counter(seq)
    # (A) 인덱스열 중복
    for c, lab in ordered:
        if _INDEX_RE.match(lab.strip()) and counts[lab] > 1:
            sbs_cols.add(c)
    # (B) ≥2개 'distinct' 라벨로 된 블록이 비겹침으로 2회 이상 반복.
    #     - distinct 조건: 매트릭스의 '같은 라벨 인접 반복'(이석영,이석영 / 박은희,박은희)을 제외.
    #     - 비겹침 2회: 두 표가 좌우로 나란히 같은 하위컬럼(시스템,방식 …)을 갖는 구조(NAC연계).
    n = len(seq)
    for L in range(2, n // 2 + 1):
        windows: Dict[tuple, List[int]] = {}
        for s in range(0, n - L + 1):
            windows.setdefault(tuple(seq[s:s + L]), []).append(s)
        for blk, starts in windows.items():
            if len(set(blk)) < 2:  # 블록 내 라벨이 모두 같으면(매트릭스 셀) 제외
                continue
            chosen: List[int] = []
            last = -1
            for s in starts:
                if s > last:           # 비겹침 점유
                    chosen.append(s)
                    last = s + L - 1
            if len(chosen) >= 2:
                for s in chosen:
                    for off in range(L):
                        sbs_cols.add(cols[s + off])
    return sbs_cols, cols, seq


def _header_labels(region, canvas) -> Dict[int, str]:
    """region 헤더행의 {col: label}. header_rows 없으면 빈 dict."""
    out: Dict[int, str] = {}
    for hr in (region.header_rows or []):
        for col in range(region.min_col, region.max_col + 1):
            cell = canvas.cells.get((hr, col))
            # CellNode 필드명: display_value / normalized_value / logical_value (cell_node.py).
            # 병합·복원 셀까지 잡으려면 logical_value 우선.
            val = "" if cell is None else ("" if cell.is_empty else str(getattr(cell, "logical_value", "") or cell.display_value or cell.normalized_value or "").strip())
            if val and col not in out:
                out[col] = val
    return out


def compute_gate_summary(input_path, chunks: List[Dict[str, Any]]) -> Dict[str, Any]:
    path = Path(input_path)
    cfg = ParserConfig()
    canvases = build_canvases(path, cfg)
    region_pairs = detect_and_classify(canvases, cfg)

    # region 을 시트별로 묶기
    by_sheet: Dict[str, list] = defaultdict(list)
    for region, canvas in region_pairs:
        by_sheet[canvas.sheet_name].append((region, canvas))

    # 원시 워크북(참조오류 스캔용).
    #  - data_only=True : 캐시된 계산 결과의 에러 문자열(#REF! 등)
    #  - data_only=False: 수식 문자열 자체의 깨진 참조(=SUM(#REF!) 등) — 캐시가 없을 때도 잡는다.
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        wb_formula = openpyxl.load_workbook(path)  # 기본=수식 보존
    except Exception:
        wb_formula = None

    sheets_out: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        findings: List[Dict[str, Any]] = []

        # 1) ref_error — 값(캐시) + 수식 문자열 양쪽 스캔
        ref_set: set = set()
        ws_f = wb_formula[ws.title] if (wb_formula is not None and ws.title in wb_formula.sheetnames) else None
        for src_ws in (ws, ws_f):
            if src_ws is None:
                continue
            for row in src_ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and ERROR_RE.search(cell.value):
                        ref_set.add((cell.row, cell.column))
        ref_cells = [f"{get_column_letter(c)}{r}" for r, c in sorted(ref_set)]
        if ref_cells:
            # 플래그 좌표마다 무조건 캐시값(#REF!)+수식(=J9)을 조회.
            #   - H3=`=J9` 처럼 에러토큰 없는 수식셀은 캐시 #REF! 로만 플래그되므로
            #     ERROR_RE 매치 분기 안에서 수식을 잡으면 도달 불가 → 좌표 무조건 조회.
            #   - MergedCell 방어: getattr 로 value 안전 접근.
            parts: List[str] = []
            for r, c in sorted(ref_set):
                coord = f"{get_column_letter(c)}{r}"
                value = getattr(ws.cell(r, c), "value", None)
                formula = getattr(ws_f.cell(r, c), "value", None) if ws_f is not None else None
                fml = str(formula)[:60] if formula is not None else None
                if value is None and fml:
                    parts.append(f"{coord}=({fml})")
                elif fml:
                    parts.append(f"{coord}={value} ({fml})")
                else:
                    parts.append(f"{coord}={value}")
                if len(parts) >= 5:
                    break
            findings.append({"code": "ref_error", "cells": ref_cells[:20],
                             "detail": f"참조 오류가 값에 포함됨: {', '.join(parts)}"})

        # 2)~3) side_by_side / empty_header — region 헤더 기반
        for region, canvas in by_sheet.get(ws.title, []):
            labels = _header_labels(region, canvas)
            if not labels:
                continue
            sbs_cols, cols, seq = _detect_side_by_side(labels)
            if sbs_cols:
                hr0 = region.header_rows[0]
                dup_cells = sorted(f"{get_column_letter(c)}{hr0}" for c in sbs_cols)
                involved = sorted({seq[cols.index(c)] for c in sbs_cols})
                hdr_seq = " | ".join(seq[:12]) + (" …" if len(seq) > 12 else "")
                findings.append({"code": "side_by_side", "cells": dup_cells,
                                 "detail": f"나란히 놓인 두 표로 판단(중복/반복 헤더: {', '.join(involved)})"
                                           f"(헤더행: {hdr_seq})"})
            # empty_header: 사용 열에 헤더 라벨이 비어있는 칸.
            # ⚠️ 의도적으로 거의 비활성(보수적): region 이 잡혔다는 건 보통 라벨 ≥2 이므로
            #   아래 `len(labels) < 2` 게이트로 실질적 미발화. trailing blank column 오탐을
            #   피하려는 스캐폴딩이며, 추후 명확한 트리거 정의가 생기면 완화한다.
            # 임계치: 전체 region 열 중 빈 헤더 비율 50% 초과 AND 라벨<2 일 때만.
            total_cols = region.max_col - region.min_col + 1
            empty_cols = [get_column_letter(col) + str(region.header_rows[0])
                          for col in range(region.min_col, region.max_col + 1) if col not in labels]
            empty_ratio = len(empty_cols) / total_cols if total_cols > 0 else 0
            if empty_cols and empty_ratio > 0.5 and len(labels) < 2:
                findings.append({"code": "empty_header", "cells": empty_cols[:20],
                                 "detail": f"헤더 컬럼명이 비어있음: {', '.join(empty_cols[:5])}"})

        # 3b) ambiguous_hierarchy — 계층열의 최소 번호레벨이 strictly 증가하지 않으면
        #     (깊은 열이 얕은 열과 같은 단계로 번호를 재시작) 상하위 관계가 모호 (SoT §계층).
        #     region_type=='hierarchical_matrix' 한정(precision 우선). cells 는 실제 번호 body 셀.
        #     [알려진 한계 — false-negative] 이 신호는 열 레벨을 item_numbering_level 로만 근사한다.
        #     대분류가 '미인식 마커'(①/i/a 등, item_numbering_level=None)인 경우, 파서는
        #     컬럼위치 fallback 으로 그 열에 level(예 0)을 부여해 하위의 인식마커('1.', level0)와
        #     충돌·붕괴하지만, 이 스캔은 그 열을 못 봐서 놓친다. 대분류가 인식 마커(1./가./I.)인
        #     흔한 케이스(예: 직무전결)는 정확히 커버. 근본 커버는 증상기반 신호(고아율 등, 후속).
        for region, canvas in by_sheet.get(ws.title, []):
            if region.region_type != "hierarchical_matrix":
                continue
            hcols = sorted(region.hierarchy_cols or [])
            if len(hcols) < 2:
                continue
            col_min_cell = {}  # c -> (row, min_level): 그 열의 최소 번호레벨을 달성한 첫 body 셀
            for r in body_rows_of(region, canvas):
                for c in hcols:
                    t = cell_text(canvas.get_cell(r, c))
                    if not t:
                        continue
                    lv = item_numbering_level(t)
                    if lv is None:
                        continue
                    prev = col_min_cell.get(c)
                    if prev is None or lv < prev[1]:
                        col_min_cell[c] = (r, lv)
            ordered = [(c, col_min_cell[c][1]) for c in hcols if c in col_min_cell]
            bad = _numbering_restart(ordered)
            if bad:
                cs, cd = bad
                findings.append({
                    "code": "ambiguous_hierarchy",
                    "cells": [f"{get_column_letter(cs)}{col_min_cell[cs][0]}",
                              f"{get_column_letter(cd)}{col_min_cell[cd][0]}"],
                    "detail": (f"깊은 계층 열({get_column_letter(cd)})의 번호가 얕은 열"
                               f"({get_column_letter(cs)})과 같은 단계로 재시작되어 상·하위 관계가 모호함"),
                })

        # 4) header_leak — chunk 의 field[k]==k (헤더행이 데이터로 추출됨)
        for c in chunks:
            if c.get("sheet") != ws.title:
                continue
            fields = c.get("fields") or {}
            if not isinstance(fields, dict) or len(fields) < 2:
                continue
            same = sum(1 for k, v in fields.items()
                       if isinstance(v, str) and v.strip() == str(k).strip() and v.strip() != "")
            if same >= max(2, (len(fields) + 1) // 2):
                src = c.get("source") or {}
                row = src.get("start_row")
                loc = [f"row{row}"] if row else []
                leaked = [str(k) for k, v in fields.items()
                          if isinstance(v, str) and v.strip() == str(k).strip() and v.strip()][:5]
                pairs = ", ".join(f"{k}={k}" for k in leaked)
                # row{n}(청크 데이터행)은 실제 헤더행과 다를 수 있어 모호 → 헤더=값 쌍으로 자기설명.
                findings.append({"code": "header_leak", "cells": loc,
                                 "detail": f"헤더행이 데이터로 추출됨(이 행이 헤더 라벨과 동일): {pairs}"})
                break  # 시트당 1건이면 충분

        sheets_out.append({"sheet": ws.title, "ok": not findings, "findings": findings})

    wb.close()
    if wb_formula is not None:
        wb_formula.close()
    return {"ok": all(s["ok"] for s in sheets_out), "sheets": sheets_out}
