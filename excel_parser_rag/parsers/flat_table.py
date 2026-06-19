"""FlatTableParser — 단순 표 파서 (SoT §10.2 flat_table, §12, §17.3, §19.2).

이 모듈은 다른 파서들이 재사용하는 공용 헬퍼도 함께 제공한다:
- cell_text / merged_text : 셀 텍스트 추출 (raw 우선, 병합 logical 보조)
- flatten_headers         : 다단 헤더 flatten (SoT §11.4)
- body_rows_of            : region 의 본문 행 목록
- region_row_text         : 행 전체 텍스트 결합
"""

from __future__ import annotations

from typing import Any, Dict, List, Tuple, TYPE_CHECKING

from openpyxl.utils import get_column_letter

from ..chunking.chunk_schema import RagChunk
from ..textutil import is_note_text, is_total_text, one_line
from .base import BaseRegionParser, ParseContext

if TYPE_CHECKING:
    from ..canvas.cell_node import CellNode
    from ..canvas.sheet_canvas import SheetCanvas
    from ..detection.region import Region


# ---------------------------------------------------------------------------
# 공용 헬퍼
# ---------------------------------------------------------------------------

def cell_text(cell: "CellNode") -> str:
    """셀의 원본(raw/display) 텍스트. 병합 전파값(logical)은 포함하지 않는다."""
    return one_line(cell.display_value) or one_line(cell.raw_value)


def merged_text(cell: "CellNode", orientations: Tuple[str, ...] = ("vertical", "block")) -> str:
    """병합 logical 보조값. 지정한 병합 방향일 때만 반환한다 (SoT §7.3)."""
    if cell.merge_orientation in orientations:
        return one_line(cell.logical_value)
    return ""


def flatten_headers(region: "Region", canvas: "SheetCanvas", use_region_cols: bool = True) -> Dict[int, str]:
    """다단 헤더 flatten (SoT §11.4) — col index -> 컬럼명.

    region.matrix_cols / metadata_cols 가 이미 채워져 있으면 그것을 우선 사용한다
    (header_detector 또는 config override 가 채움).
    """
    if use_region_cols and (region.matrix_cols or region.metadata_cols):
        names: Dict[int, str] = {int(c): str(n) for c, n in region.matrix_cols.items()}
        for c, n in region.metadata_cols.items():
            names.setdefault(int(c), str(n))
        if names:
            return names

    names = {}
    used: Dict[str, int] = {}
    for c in range(region.min_col, region.max_col + 1):
        parts: List[str] = []
        for r in region.header_rows:
            cell = canvas.get_cell(r, c)
            value = one_line(cell.logical_value) or cell_text(cell)
            if value and (not parts or parts[-1] != value):
                parts.append(value)
        name = "_".join(parts)
        if not name:
            continue
        if name in used:  # 같은 이름 충돌 시 suffix (SoT §11.4)
            used[name] += 1
            name = f"{name}_{used[name]}"
        else:
            used[name] = 1
        names[c] = name
    return names


def body_rows_of(region: "Region", canvas: "SheetCanvas") -> List[int]:
    """region 의 본문 행. region.body_rows 가 있으면 그대로, 없으면 헤더 아래 전체."""
    if region.body_rows:
        rows = [r for r in region.body_rows if region.min_row <= r <= region.max_row]
    else:
        start = (max(region.header_rows) + 1) if region.header_rows else region.min_row
        footers = set(region.footer_rows)
        rows = [r for r in range(start, region.max_row + 1) if r not in footers]
    headers = set(region.header_rows)
    return [r for r in rows if r not in headers]


def region_row_text(canvas: "SheetCanvas", region: "Region", row: int) -> str:
    parts = []
    for c in range(region.min_col, region.max_col + 1):
        t = cell_text(canvas.get_cell(row, c))
        if t:
            parts.append(t)
    return " ".join(parts)


# ---------------------------------------------------------------------------
# FlatTableParser
# ---------------------------------------------------------------------------

class FlatTableParser(BaseRegionParser):
    """단순 표: 행마다 table_row chunk + total_row 분리 + table_summary 1개."""

    name = "flat_table"
    region_types = ("flat_table", "report_section", "unknown_table")

    row_confidence = 0.88
    total_confidence = 0.85
    summary_confidence = 0.9
    note_confidence = 0.82

    def parse(self, region: "Region", canvas: "SheetCanvas", ctx: ParseContext) -> List[RagChunk]:
        headers = flatten_headers(region, canvas)
        chunks: List[RagChunk] = []
        data_rows = 0
        note_count = 0

        for r in body_rows_of(region, canvas):
            if not canvas.row_has_content(r, region.min_col, region.max_col):
                continue
            fields, first_label = self._row_fields(region, canvas, r, headers)
            if not fields:
                continue

            joined = " ".join(str(v) for v in fields.values())
            if is_note_text(first_label):
                chunks.append(self._note_chunk(region, canvas, ctx, r, joined))
                note_count += 1
                continue

            is_total = any(is_total_text(v) for v in list(fields.values())[:2])
            chunk_type = "total_row" if is_total else "table_row"
            chunk = self.new_chunk(region, canvas, ctx, chunk_type, min_row=r, max_row=r)
            chunk.path = [p for p in (chunk.title, first_label) if p]
            chunk.fields = dict(fields)
            chunk.facts = [{"predicate": k, "value": v} for k, v in fields.items()]
            sentences = ", ".join(f"{k}는 {v}" for k, v in fields.items())
            chunk.content_text = (
                f"{ctx.document_title}의 {canvas.sheet_name} 시트에서 "
                f"'{' > '.join(chunk.path)}' 항목은 다음 값을 가진다: {sentences}. "
                f"원본 위치는 {chunk.range}이다."
            )
            chunk.metadata["is_total"] = is_total
            chunk.quality = {"confidence": self.total_confidence if is_total else self.row_confidence}
            if not is_total:
                data_rows += 1
            chunks.append(chunk)

        footer_chunks = self._footer_chunks(region, canvas, ctx)
        note_count += sum(1 for c in footer_chunks if c.chunk_type == "note")
        chunks.extend(footer_chunks)

        chunks.insert(0, self._table_summary(region, canvas, ctx, headers, data_rows, note_count))
        return chunks

    # --- 내부 ---------------------------------------------------------------

    def _row_fields(
        self, region: "Region", canvas: "SheetCanvas", row: int, headers: Dict[int, str]
    ) -> Tuple[Dict[str, Any], str]:
        fields: Dict[str, Any] = {}
        first_label = ""
        for c in range(region.min_col, region.max_col + 1):
            cell = canvas.get_cell(row, c)
            value = cell_text(cell) or merged_text(cell, ("vertical",))
            if not value:
                continue
            key = headers.get(c) or get_column_letter(c)
            if key not in fields:
                fields[key] = value
            if not first_label:
                first_label = value
        return fields, first_label

    def _note_chunk(
        self, region: "Region", canvas: "SheetCanvas", ctx: ParseContext, row: int, text: str
    ) -> RagChunk:
        chunk = self.new_chunk(region, canvas, ctx, "note", min_row=row, max_row=row)
        related = chunk.title or ctx.document_title
        chunk.path = [related] if related else []
        chunk.fields = {"주석": text}
        chunk.content_text = f"{ctx.document_title}의 {related} 관련 주석: {text}"
        chunk.quality = {"confidence": self.note_confidence}
        return chunk

    def _footer_chunks(self, region: "Region", canvas: "SheetCanvas", ctx: ParseContext) -> List[RagChunk]:
        out: List[RagChunk] = []
        for r in region.footer_rows:
            text = region_row_text(canvas, region, r)
            if not text:
                continue
            if is_total_text(text.split()[0] if text.split() else ""):
                chunk = self.new_chunk(region, canvas, ctx, "total_row", min_row=r, max_row=r)
                chunk.path = [p for p in (chunk.title,) if p]
                chunk.fields = {"합계행": text}
                chunk.content_text = (
                    f"{ctx.document_title}의 {canvas.sheet_name} 시트 '{chunk.title}' 표의 합계 행: {text}"
                )
                chunk.metadata["is_total"] = True
                chunk.quality = {"confidence": self.total_confidence}
                out.append(chunk)
            else:
                out.append(self._note_chunk(region, canvas, ctx, r, text))
        return out

    def _table_summary(
        self,
        region: "Region",
        canvas: "SheetCanvas",
        ctx: ParseContext,
        headers: Dict[int, str],
        data_rows: int,
        note_count: int,
    ) -> RagChunk:
        summary = self.new_chunk(region, canvas, ctx, "table_summary")
        cols = [headers[c] for c in sorted(headers)][:12]
        summary.path = [summary.title] if summary.title else []
        summary.fields = {
            "범위": region.range_a1,
            "데이터행수": data_rows,
            "주석수": note_count,
            "컬럼": cols,
        }
        col_text = f" 컬럼은 {', '.join(cols)}이다." if cols else ""
        summary.content_text = (
            f"{ctx.document_title}의 {canvas.sheet_name} 시트에 있는 '{summary.title}' 표는 "
            f"총 {data_rows}개의 데이터 행을 가진다.{col_text}"
        )
        summary.quality = {"confidence": self.summary_confidence}
        return summary
