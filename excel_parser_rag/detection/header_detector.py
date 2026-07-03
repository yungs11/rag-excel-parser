"""Header 탐지 (SoT §11) + 컬럼 역할 부여.

detect_headers(region, canvas, config):
- header_score 로 상단 연속 헤더 행 판정 (max_header_depth)
- 다단 헤더 flatten → 컬럼명 생성
  - 빈 헤더는 위/왼쪽 logical 로 보강 (병합은 logical_value 가 이미 처리)
  - matrix 계열 region 은 "전결권자" 같은 상위 그룹 헤더 대신 하위 라벨 우선
  - 이름 충돌 시 suffix
- region.header_rows / hierarchy_cols / matrix_cols / metadata_cols / body_rows 채움
- override 로 이미 채워진 값은 유지 (빈 이름만 보강)
"""

from __future__ import annotations

import re
from typing import Dict, List, Optional, Set, Tuple, TYPE_CHECKING

from openpyxl.utils import range_boundaries

from ..config import ParserConfig
from ..markerutil import is_marker_cell
from ..textutil import compact, infer_numbering_level, one_line

if TYPE_CHECKING:
    from ..canvas.cell_node import CellNode
    from ..canvas.sheet_canvas import SheetCanvas
    from ..detection.region import Region

_HEADER_SCORE_MIN = 1.8
_STYLE_GATE_MIN = 0.5   # bold/배경색/가운데정렬 등 스타일 신호 최소값
_MAX_PRE_ROWS = 8       # 헤더 시작 전 스킵 가능한 희소 메타 행 수

# 합의/수신/비고 류 — metadata_cols 로 분리할 헤더명 (compact 비교)
_METADATA_HEADER_TERMS = {"합의", "수신", "비고", "참고", "참조", "근거", "관련근거", "관련규정"}
# 항목/계층 컬럼 헤더로 보는 어휘 (compact 후 부분 일치)
_ITEM_HEADER_TERMS = ("사항", "항목", "구분", "내용", "업무", "분류", "제목", "품목", "품명")

# 헤더 탐지를 건너뛰는 region 유형 (key-value/주석/제목 류)
_HEADERLESS_TYPES = {"code_mapping_table", "form", "key_value_block", "note_block", "report_section"}

_NUMERIC_RE = re.compile(r"^-?[\d,]+(?:\.\d+)?%?$")


# --- 공용 헬퍼 -------------------------------------------------------------------

def _cell_text(cell: Optional["CellNode"]) -> str:
    if cell is None:
        return ""
    v = cell.logical_value or cell.normalized_value or cell.display_value
    if not v and cell.raw_value is not None:
        v = cell.raw_value
    return one_line(v)


def _is_numeric(cell: Optional["CellNode"], text: str) -> bool:
    if cell is not None and cell.data_type in ("int", "float"):
        return True
    return bool(_NUMERIC_RE.match(compact(text)))


def _merge_span_cols(cell: "CellNode", canvas: "SheetCanvas") -> Tuple[int, int]:
    rng = cell.merge_range
    if not rng:
        for mr in canvas.merged_ranges:
            try:
                c0, r0, c1, r1 = range_boundaries(mr)
            except Exception:
                continue
            if r0 <= cell.row <= r1 and c0 <= cell.col <= c1:
                return int(c0), int(c1)
        return cell.col, cell.col
    try:
        c0, _r0, c1, _r1 = range_boundaries(rng)
        return int(c0), int(c1)
    except Exception:
        return cell.col, cell.col


def _is_merge_shadow_row(canvas: "SheetCanvas", row: int, header_rows: List[int]) -> bool:
    """row 의 셀 중 하나라도 header_rows 에서 시작하는 세로/블록 병합에 덮이면 True.

    스팬 부모헤더(전결권자 F1:G1) 아래의 무스타일·희소 leaf 행(부총장/처장)을
    헤더밴드로 편입하기 위한 신호. 세로병합(부서명 A1:A2 등)이 헤더행에서 시작해
    아래로 뻗으면 그 밑 행도 같은 헤더 구조의 일부다.
    """
    hset = set(header_rows)
    for mr in canvas.merged_ranges:
        try:
            _c0, r0, _c1, r1 = range_boundaries(mr)
        except Exception:
            continue
        if r0 in hset and r0 < row <= r1:
            return True
    return False


def _title_rows(region: "Region") -> Set[int]:
    """region.title_range 가 region 내부에 있으면 해당 행들을 반환."""
    if not region.title_range:
        return set()
    try:
        _c0, r0, _c1, r1 = range_boundaries(region.title_range)
    except Exception:
        return set()
    rows = {r for r in range(int(r0), int(r1) + 1) if region.min_row <= r <= region.max_row}
    return rows


# --- row metrics / header score (SoT §11.2) --------------------------------------

def _row_metrics(canvas: "SheetCanvas", row: int, cols: List[int]) -> Dict[str, float]:
    n = len(cols)
    filled = bold = center = fillc = border = nonnum = short = hmerge = marker = numbering = numeric = 0
    for c in cols:
        cell = canvas.cells.get((row, c))
        text = _cell_text(cell)
        if not text:
            continue
        filled += 1
        st = cell.style
        if st.bold:
            bold += 1
        if st.horizontal_alignment in ("center", "centerContinuous"):
            center += 1
        if st.fill_color:
            fillc += 1
        if st.border_top or st.border_bottom or st.border_left or st.border_right:
            border += 1
        if cell.merge_orientation in ("horizontal", "block"):
            hmerge += 1
        if is_marker_cell(cell, text):
            marker += 1
        if infer_numbering_level(text) is not None:
            numbering += 1
        if _is_numeric(cell, text):
            numeric += 1
        else:
            nonnum += 1
        if len(text) <= 14:
            short += 1
    if filled == 0:
        return {"filled": 0.0, "style_signal": 0.0}
    f = float(filled)
    m = {
        "filled": float(filled),
        "bold": bold / f,
        "center": center / f,
        "fillcolor": fillc / f,
        "border": border / f,
        "text": nonnum / f,
        "short": short / f,
        "hmerge": hmerge / f,
        "fill": filled / float(n),
        "marker_ratio": marker / f,
        "numbering_ratio": numbering / f,
        "numeric": numeric / f,
    }
    m["style_signal"] = m["bold"] + m["fillcolor"] + (0.5 if m["center"] >= 0.5 else 0.0)
    return m


def _header_score(m: Dict[str, float]) -> float:
    """SoT §11.2 — bold/fill/border/text/short/merge 가산, marker/번호/숫자 감점."""
    return (
        1.2 * m["bold"]
        + 0.8 * m["center"]
        + 0.6 * m["fillcolor"]
        + 0.4 * m["border"]
        + 0.8 * m["text"]
        + 0.5 * m["short"]
        + 0.4 * m["hmerge"]
        + 0.6 * m["fill"]
        - 1.5 * m["marker_ratio"]
        - 1.5 * m["numbering_ratio"]
        - 0.8 * m["numeric"]
    )


def _detect_header_rows(
    region: "Region",
    canvas: "SheetCanvas",
    cols: List[int],
    title_rows: Set[int],
    max_depth: int,
) -> Tuple[List[int], Set[int], float]:
    """상단에서 연속 header 행 탐지. (header_rows, pre_rows(헤더 위 메타행), best_score)"""
    pre: Set[int] = set()
    headers: List[int] = []
    best = 0.0

    r = region.min_row
    while r <= region.max_row:
        if r in title_rows:
            if headers:
                break
            pre.add(r)
            r += 1
            continue
        m = _row_metrics(canvas, r, cols)
        if m["filled"] == 0:
            if headers:
                break
            pre.add(r)
            r += 1
            continue
        if m.get("numbering_ratio", 0.0) > 0 or m.get("marker_ratio", 0.0) > 0:
            break  # 본문 시작 (항목 번호/마커 등장)
        score = _header_score(m)
        strong = (
            score >= _HEADER_SCORE_MIN
            and m["style_signal"] >= _STYLE_GATE_MIN
            and m["filled"] >= 2
        )
        # 스팬 부모헤더 아래 leaf 행: 무스타일·희소라 style gate 는 탈락하지만,
        # 헤더행에서 시작한 세로/블록 병합에 덮이면(=헤더밴드 연장) 헤더로 편입한다.
        # 마커/번호 행은 위에서 이미 break 되므로 본문행 오편입 위험 없음.
        shadow_cont = (
            not strong
            and bool(headers)
            and m["filled"] >= 2
            and score >= _HEADER_SCORE_MIN
            and _is_merge_shadow_row(canvas, r, headers)
        )
        if strong or shadow_cont:
            headers.append(r)
            best = max(best, score)
            # shadow leaf 행(스팬 부모 아래 sub-header)은 헤더밴드의 끝이다.
            # 그 아래는 본문이므로 더 진행하지 않는다(블록병합·스타일된 데이터
            # 카테고리 행이 strong 으로 오편입되는 것을 차단).
            if shadow_cont or len(headers) >= max_depth:
                break
        else:
            if headers:
                break
            # 헤더 시작 전 메타 행 ("<별표N>", "(개정 : ...)" 단독 셀 등) — 한도 내에서 스킵
            pre.add(r)
            if len(pre) >= _MAX_PRE_ROWS:
                break
        r += 1
    return headers, pre, best


# --- multi-row header flatten (SoT §11.4) -----------------------------------------

def _collect_header_parts(
    canvas: "SheetCanvas", cols: List[int], header_rows: List[int]
) -> Dict[int, List[str]]:
    parts_by_col: Dict[int, List[str]] = {}
    for c in cols:
        parts: List[str] = []
        for r in header_rows:
            v = _cell_text(canvas.cells.get((r, c)) or canvas.get_cell(r, c))
            if v and (not parts or parts[-1] != v):
                parts.append(v)
        parts_by_col[c] = parts
    return parts_by_col


def _flatten_column_names(
    region: "Region", cols: List[int], parts_by_col: Dict[int, List[str]]
) -> Dict[int, str]:
    matrix_mode = region.region_type in ("matrix_table", "hierarchical_matrix")
    names: Dict[int, str] = {}
    used: Dict[str, int] = {}
    prev_name = ""
    for c in cols:
        parts = parts_by_col.get(c, [])
        if parts:
            # matrix 계열: 상위 그룹 헤더("전결권자")보다 하위 라벨("팀장") 우선
            name = parts[-1] if matrix_mode else "_".join(parts)
            prev_name = name
        else:
            # 빈 헤더 — 왼쪽 logical 보강 (왼쪽도 없으면 "")
            name = prev_name
        if name:
            used[name] = used.get(name, 0) + 1
            if used[name] > 1:
                name = f"{name}_{used[name]}"
        names[c] = name
    return names


# --- 컬럼 역할 부여 ----------------------------------------------------------------

def _detect_hierarchy_cols(
    region: "Region",
    canvas: "SheetCanvas",
    cols: List[int],
    parts_by_col: Dict[int, List[str]],
    header_rows: List[int],
) -> List[int]:
    colset = set(cols)
    first_named: Optional[int] = None
    for c in cols:
        if parts_by_col.get(c):
            first_named = c
            break

    if header_rows and first_named is not None:
        anchor = canvas.get_cell(header_rows[0], first_named)
        c0, c1 = _merge_span_cols(anchor, canvas)
        if c1 > c0:
            # "전 결 사 항" 처럼 항목 헤더가 가로(또는 블록) 병합으로 여러 컬럼을 차지
            return [c for c in range(c0, c1 + 1) if c in colset]
        label = compact(parts_by_col[first_named][0])
        if any(term in label for term in _ITEM_HEADER_TERMS):
            return [first_named]

    # fallback: 왼쪽 컬럼들의 항목 번호 패턴 스캔
    body_start = (max(header_rows) + 1) if header_rows else region.min_row
    hier: List[int] = []
    for c in cols[: min(5, len(cols))]:
        texts: List[str] = []
        for r in range(body_start, region.max_row + 1):
            cell = canvas.cells.get((r, c))
            if cell is None or cell.is_empty:
                continue
            t = _cell_text(cell)
            if t and not is_marker_cell(cell, t):
                texts.append(t)
        if len(texts) < 3:
            continue
        hits = sum(1 for t in texts if infer_numbering_level(t) is not None)
        if hits / len(texts) >= 0.3:
            hier.append(c)
    if hier:
        return hier
    if region.region_type in ("matrix_table", "hierarchical_matrix") and first_named is not None:
        return [first_named]
    return []


def _assign_column_roles(
    region: "Region",
    canvas: "SheetCanvas",
    cols: List[int],
    names: Dict[int, str],
    parts_by_col: Dict[int, List[str]],
    header_rows: List[int],
) -> None:
    if not region.hierarchy_cols:
        region.hierarchy_cols = _detect_hierarchy_cols(region, canvas, cols, parts_by_col, header_rows)
    hier = set(region.hierarchy_cols)

    if region.metadata_cols:
        for c, nm in list(region.metadata_cols.items()):
            if not nm:
                region.metadata_cols[c] = names.get(c, "")
    else:
        for c in cols:
            if c in hier:
                continue
            nm = names.get(c, "")
            if nm and compact(nm) in _METADATA_HEADER_TERMS:
                region.metadata_cols[c] = nm

    if region.matrix_cols:
        for c, nm in list(region.matrix_cols.items()):
            if not nm:
                region.matrix_cols[c] = names.get(c, "")
    else:
        for c in cols:
            if c in hier or c in region.metadata_cols:
                continue
            nm = names.get(c, "")
            if not nm:
                continue  # 이름 없는 컬럼(예: 좌측 순번 열)은 어디에도 배정하지 않음
            region.matrix_cols[c] = nm


# --- 엔트리 포인트 ------------------------------------------------------------------

def detect_headers(region: "Region", canvas: "SheetCanvas", config: ParserConfig) -> None:
    cols = list(range(region.min_col, region.max_col + 1))
    title_rows = _title_rows(region)

    def content_rows(start: int, excluded: Set[int]) -> List[int]:
        return [
            r
            for r in range(start, region.max_row + 1)
            if r not in excluded and canvas.row_has_content(r, region.min_col, region.max_col)
        ]

    # key-value/주석/제목 류 region 은 헤더 탐지를 건너뛴다
    if region.region_type in _HEADERLESS_TYPES or region.role != "body":
        if not region.body_rows:
            region.body_rows = content_rows(region.min_row, title_rows)
        return

    if region.header_rows:
        # override 등으로 이미 지정 — 유지
        header_rows = sorted(region.header_rows)
        pre_rows: Set[int] = set(range(region.min_row, header_rows[0]))
    else:
        header_rows, pre_rows, best = _detect_header_rows(
            region, canvas, cols, title_rows, max(1, int(config.max_header_depth))
        )
        region.header_rows = header_rows
        region.features["header_detection_score"] = float(best)
        if not header_rows:
            region.warnings.append("header_not_detected")

    parts_by_col = _collect_header_parts(canvas, cols, header_rows)
    names = _flatten_column_names(region, cols, parts_by_col)
    _assign_column_roles(region, canvas, cols, names, parts_by_col, header_rows)

    if not region.body_rows:
        if header_rows:
            excluded = title_rows | set(header_rows) | set(pre_rows)
            region.body_rows = content_rows(max(header_rows) + 1, excluded)
        else:
            region.body_rows = content_rows(region.min_row, title_rows)
