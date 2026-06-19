"""kordoc(.md) 기반 백엔드 (kordoc 통합설계).

흐름: xlsx → (kordoc .md 확보: 지정/디렉토리/자동생성) → HTML grid 복원 → 시트제목·배너
분할 → 섹션별 헤더 → 행 청크(compact matrix + table_row) → k2o 좌표 → SoT 호환 chunk dict.

md 확보 우선순위:
  1. config.kordoc_md_path (단일 파일)
  2. config.kordoc_md_dir/<stem>.md
  3. config.kordoc_bin 으로 자동 생성 (Node 필요) → config.kordoc_md_out (기본 임시)
없으면 명확히 에러(자동 openpyxl fallback 금지 — silent 품질저하 방지).
"""
from __future__ import annotations

import difflib
import os
import re
import shlex
import subprocess
import tempfile
from collections import Counter
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..config import ParserConfig
from ..textutil import PARSER_VERSION, infer_numbering_level
from .base import BackendError

# ─── 헤더 스코어러 상수 ──────────────────────────────
_MAX_HEADER_SCAN = 8
_NUMERIC_RE = re.compile(r"^-?[\d,]+(?:\.\d+)?%?$")
_FORMULA_ERR_RE = re.compile(r"#(REF|VALUE|DIV/0|N/A|NAME\?|NULL|NUM)!?")

# ─── 마커 ────────────────────────────────────────────
MARKER_NORM = {
    "○": "applicable", "◯": "applicable", "●": "applicable_primary", "◎": "applicable_special",
    "△": "conditional", "▲": "conditional", "×": "not_applicable", "✕": "not_applicable",
    "✓": "checked", "✔": "checked", "√": "checked",
}
MARK_BUCKET = {
    "○": "해당", "◯": "해당", "●": "해당", "◎": "해당", "△": "조건부", "▲": "조건부",
    "×": "비해당", "✕": "비해당", "✓": "체크", "✔": "체크", "√": "체크",
}


def is_marker(t: Any) -> bool:
    return str(t or "").strip() in MARKER_NORM


def clean_title(t: Any) -> str:
    return re.sub(r"^[▶▷●■◆\s]+", "", str(t or "")).strip()


def clean_val(t: Any) -> str:
    return re.sub(r"\s*\n\s*", " ", str(t)).strip()


def norm(s: Any) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


# ─── kordoc .md → grid ───────────────────────────────
class _TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rawrows: List[List[Tuple[str, int, int, str]]] = []
        self._row = None
        self._buf: List[str] = []
        self._cs = self._rs = 1
        self._tag = "td"

    def handle_starttag(self, tag, attrs):
        a = dict(attrs)
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th"):
            self._buf = []
            self._cs = int(a.get("colspan", 1) or 1)
            self._rs = int(a.get("rowspan", 1) or 1)
            self._tag = tag
        elif tag == "br":
            self._buf.append("\n")

    def handle_data(self, d):
        if self._row is not None:
            self._buf.append(d)

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._row is not None:
            self._row.append(("".join(self._buf).strip(), self._cs, self._rs, self._tag))
        elif tag == "tr" and self._row is not None:
            self.rawrows.append(self._row)
            self._row = None


def _expand(rawrows):
    anchors, covered, occupied = {}, {}, set()
    ncols = 0
    for r, row in enumerate(rawrows):
        col = 1
        for (text, cs, rs, tag) in row:
            while (r, col) in occupied:
                col += 1
            for dr in range(rs):
                for dc in range(cs):
                    occupied.add((r + dr, col + dc))
                    covered[(r + dr, col + dc)] = (r, col)
            anchors[(r, col)] = (text, cs, rs, tag)
            col += cs
            ncols = max(ncols, col - 1)
    return anchors, covered, len(rawrows), ncols


def _parse_md_table(lines):
    rows = []
    for ln in lines:
        s = ln.strip()
        if not (s.startswith("|") and s.endswith("|")):
            continue
        cells = [c.strip() for c in s.strip("|").split("|")]
        if cells and set("".join(cells).replace("-", "").replace(":", "").strip()) == set():
            continue
        rows.append([(c, 1, 1, "td") for c in cells])
    return rows


def _split_sheets(md):
    out, cur, body = [], None, []
    for ln in md.splitlines():
        m = re.match(r"^##\s+(.*)$", ln)
        if m:
            if cur is not None:
                out.append((cur, body))
            cur, body = m.group(1).strip(), []
        elif cur is not None:
            body.append(ln)
    if cur is not None:
        out.append((cur, body))
    return out


def _grid(body_lines):
    text = "\n".join(body_lines)
    if "<table" in text:
        tp = _TableParser()
        tp.feed(text)
        return _expand(tp.rawrows)
    return _expand(_parse_md_table(body_lines))


# ─── 원본 좌표 (k2o) ─────────────────────────────────
def _original_eff(ws):
    covered = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                covered[(r, c)] = (rng.min_row, rng.min_col)
    eff = {}
    for row in ws.iter_rows():
        for cell in row:
            eff[(cell.row, cell.column)] = cell.value

    def effective(r, c):
        a = covered.get((r, c), (r, c))
        return eff.get(a)
    return effective


def _align_rows(anchors, covered, nrows, ncols, ws):
    oeff = _original_eff(ws)

    def keff(r, c):
        a = covered.get((r, c))
        return anchors[a][0] if a in anchors else None

    k_sig = ["".join(norm(keff(r, c)) for c in range(1, ncols + 1)) for r in range(nrows)]
    o_sig = ["".join(norm(oeff(o, c)) for c in range(1, ncols + 1)) for o in range(1, ws.max_row + 1)]
    k2o = {}
    for tag, i1, i2, j1, j2 in difflib.SequenceMatcher(None, k_sig, o_sig, autojunk=False).get_opcodes():
        if tag in ("equal", "replace"):
            for di in range(min(i2 - i1, j2 - j1)):
                k2o[i1 + di] = j1 + di + 1
    return k2o


# ─── 섹션 분할 ───────────────────────────────────────
def _row_cells(anchors, nrows):
    rc: Dict[int, List] = {}
    for (r, c), (text, cs, rs, tag) in anchors.items():
        rc.setdefault(r, []).append((c, text, cs, rs, tag))
    for r in rc:
        rc[r].sort()
    return rc


def _is_marker_row(nonempty) -> bool:
    return any(is_marker(t) for (_c, t, _cs, _rs, _tag) in nonempty)


def _multiheader_maps(group_cells, detail_cells):
    """2단 헤더 → (leaf_map{col:상세라벨}, group_map{col:그룹라벨}).

    group_cells(상위 스팬 헤더)와 detail_cells(아래 상세 헤더)를 합쳐 각 열의 라벨을 만든다.
    leaf 는 matrix 라벨(예: 팀장)에 그대로 쓰고, group 은 flat 행에서 접두사(업무현황_시스템구분)로 쓴다.
    """
    group_cov: Dict[int, str] = {}
    for (c, t, cs, _rs, _tag) in group_cells:
        lab = t.replace("\n", "").strip()
        if not lab:
            continue
        for col in range(c, c + max(1, cs)):
            group_cov[col] = lab
    leaf: Dict[int, str] = {}
    for (c, t, _cs, _rs, _tag) in detail_cells:
        lab = t.replace("\n", "").strip()
        if lab:
            leaf[c] = lab
    for col, g in group_cov.items():                 # 상세 라벨 없는 열은 그룹 라벨로 대체
        leaf.setdefault(col, g)
    group_map = {col: g for col, g in group_cov.items() if leaf.get(col) != g}
    return leaf, group_map


def _looks_multiheader(group_ne, detail_ne) -> bool:
    """group_ne(헤더 후보행)에 colspan 스팬이 있고 detail_ne(다음 비어있지 않은 행)이 그 아래
    상세 헤더로 보이면 True. 단층 헤더(위임전결: 전결사항 colspan4 + 아래는 ○ 데이터) 오탐을
    막기 위해 엄격히 판정한다."""
    if not detail_ne:
        return False
    if not any(cs >= 2 for (_c, _t, cs, _rs, _tag) in group_ne):
        return False
    if _is_marker_row(detail_ne):                    # 다음 행이 마커(○ 등) 데이터면 헤더 아님
        return False
    if len(detail_ne) <= len(group_ne):              # 상세 헤더는 더 잘게 나뉜다
        return False
    single = sum(1 for (_c, _t, cs, _rs, _tag) in detail_ne if cs == 1)
    if single / max(1, len(detail_ne)) < 0.7:        # 상세행은 대부분 1칸
        return False
    span_cols = set()
    for (c, _t, cs, _rs, _tag) in group_ne:
        if cs >= 2:
            span_cols.update(range(c, c + cs))
    detail_cols = {c for (c, _t, _cs, _rs, _tag) in detail_ne}
    return bool(detail_cols & span_cols)             # 상세행이 스팬 구간을 채워야 함


def _longest_contiguous(cols) -> int:
    if not cols:
        return 0
    best = run = 1
    for a, b in zip(cols, cols[1:]):
        run = run + 1 if b == a + 1 else 1
        best = max(best, run)
    return best


def _fill_vertical(r, cells, anchors, covered, band):
    """세로병합(rowspan) 값을 하위 데이터 행에 채운다. band 안의 열만, 위에서 내려온 것(ar<r)만.
    가로 colspan 전파(ar==r)는 제외."""
    if band is None:
        return cells
    present = {c for (c, *_x) in cells}
    out = list(cells)
    for col in range(band[0], band[1] + 1):
        if col in present:
            continue
        a = covered.get((r, col))
        if a is None or a not in anchors:
            continue
        ar, _ac = a
        if ar < r:
            text, _cs, _rs, tag = anchors[a]
            if text.strip():
                out.append((col, text, 1, 1, tag))
    out.sort()
    return out


def _header_band(cells):
    """헤더 행의 colspan-반영 커버 열들 중 최장 연속 구간 (lo, hi). 갭 너머 범례셀 배제용."""
    covered = set()
    for (c, t, cs, _rs, _tag) in cells:
        if not t.strip():
            continue
        for col in range(c, c + max(1, cs)):
            covered.add(col)
    if not covered:
        return None
    cols = sorted(covered)
    best_lo = best_hi = run_lo = cols[0]
    for a, b in zip(cols, cols[1:]):
        if b != a + 1:
            if a - run_lo > best_hi - best_lo:
                best_lo, best_hi = run_lo, a
            run_lo = b
    if cols[-1] - run_lo > best_hi - best_lo:
        best_lo, best_hi = run_lo, cols[-1]
    return (best_lo, best_hi)


def _is_body_row(cells) -> bool:
    """항목번호(1./가./1.1.1) 또는 마커(○ 등)가 있으면 본문 시작 신호."""
    for (_c, t, _cs, _rs, _tag) in cells:
        s = t.strip()
        if not s:
            continue
        if is_marker(s) or infer_numbering_level(s) is not None:
            return True
    return False


def _row_header_score(cells, ncols) -> float:
    """스타일-free 헤더 점수. 텍스트/짧은라벨/연속열밴드/밀도 가산, 숫자/수식오류/번호/마커 감점."""
    filled = len(cells)
    if filled == 0:
        return -1e9
    nonnum = short = numeric = ferr = numbering = marker = 0
    for (_c, t, _cs, _rs, _tag) in cells:
        txt = t.replace("\n", " ").strip()
        compact = re.sub(r"\s+", "", txt)
        is_num = bool(_NUMERIC_RE.match(compact))
        is_ferr = bool(_FORMULA_ERR_RE.search(txt))
        is_mark = is_marker(txt)
        is_numb = infer_numbering_level(txt) is not None
        if is_num:
            numeric += 1
        if is_ferr:
            ferr += 1
        if is_mark:
            marker += 1
        if is_numb:
            numbering += 1
        if not (is_num or is_ferr):
            nonnum += 1
        if len(txt) <= 14:
            short += 1
    f = float(filled)
    cols = sorted(c for (c, _t, _cs, _rs, _tag) in cells)
    contig = _longest_contiguous(cols) / max(1, ncols)
    density = filled / max(1, (cols[-1] - cols[0] + 1))
    return (
        0.8 * (nonnum / f) + 0.5 * (short / f) + 0.6 * contig + 0.4 * density
        - 0.8 * (numeric / f) - 1.0 * (ferr / f) - 1.5 * (numbering / f) - 1.5 * (marker / f)
    )


def _pick_header(window):
    """window: 연속 헤더 후보 [(idx, r, cells), …] (single_full/본문행 제외, len>=2).
    반환 (header_idx, group_idx|None). 다단헤더(그룹행+상세행) 쌍 우선, 없으면 점수 최고행."""
    if not window:
        return None, None
    for k in range(len(window) - 1):
        _gi, _gr, gcells = window[k]
        di, _dr, dcells = window[k + 1]
        if _looks_multiheader(gcells, dcells):
            return di, window[k][0]
    ncols_eff = max((max(c for (c, *_x) in cells) for (_i, _r, cells) in window), default=1)
    best = max(window, key=lambda w: _row_header_score(w[2], ncols_eff))
    return best[0], None


def _segment(anchors, covered, nrows, ncols):
    rc = _row_cells(anchors, nrows)
    seq = []
    for r in range(nrows):
        nonempty = [(c, t, cs, rs, tag) for (c, t, cs, rs, tag) in rc.get(r, []) if t.strip()]
        if nonempty:
            seq.append((r, nonempty))
    sheet_title = None
    sections: List[Dict[str, Any]] = []
    cur = None
    i = 0
    while i < len(seq):
        r, nonempty = seq[i]
        single_full = len(nonempty) == 1 and nonempty[0][2] >= max(2, ncols - 1)
        is_th = any(tag == "th" for (_, _, _, _, tag) in nonempty)
        if sheet_title is None and is_th and single_full:
            sheet_title = clean_title(nonempty[0][1])
            i += 1
            continue
        if single_full:
            title = clean_title(nonempty[0][1])
            if cur is not None and cur["header"] is None and not cur["rows"]:
                cur["title"] = (cur["title"] + " / " + title) if cur["title"] else title
            else:
                cur = {"title": title, "header": None, "header_group": {}, "header_band": None, "rows": []}
                sections.append(cur)
            i += 1
            continue
        if cur is None:
            cur = {"title": None, "header": None, "header_group": {}, "header_band": None, "rows": []}
            sections.append(cur)
        if cur["header"] is None and len(nonempty) >= 2 and not _is_body_row(nonempty):
            # 헤더 후보 윈도우: i 부터 본문시작/제목 전까지(연속), 최대 _MAX_HEADER_SCAN
            window = []
            j = i
            while j < len(seq) and len(window) < _MAX_HEADER_SCAN:
                rj, cellsj = seq[j]
                sfj = len(cellsj) == 1 and cellsj[0][2] >= max(2, ncols - 1)
                if sfj or _is_body_row(cellsj):
                    break
                if len(cellsj) >= 2:
                    window.append((j, rj, cellsj))
                j += 1
            hidx, gidx = _pick_header(window)
            if hidx is not None:
                hcells = seq[hidx][1]
                cur["header_band"] = _header_band(hcells)
                if gidx is not None:
                    leaf, gmap = _multiheader_maps(seq[gidx][1], hcells)
                    cur["header"] = leaf
                    cur["header_group"] = gmap
                else:
                    cur["header"] = {c: t.replace("\n", "").strip() for (c, t, cs, rs, tag) in hcells}
                i = hidx + 1                          # 헤더 위 메타/제목/범례 행은 drop
                continue
        filled = _fill_vertical(r, nonempty, anchors, covered, cur.get("header_band"))
        cur["rows"].append((r, filled))
        i += 1
    return sheet_title, sections


# ─── 청크 빌드 ───────────────────────────────────────
def _kw(*texts):
    toks, seen, out = [], set(), []
    for t in texts:
        for part in re.split(r"[\s/,·>()\[\]:;]+", str(t)):
            p = part.strip()
            if len(p) >= 2 and not is_marker(p):
                toks.append(p)
    for t in toks:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out[:30]


def _quality(conf, review, warns):
    return {"confidence": round(conf, 2), "review_required": bool(review),
            "warnings": warns, "parser_version": PARSER_VERSION}


def _build(doc, sheet, sheet_title, anchors, covered, nrows, ncols, k2o):
    title = sheet_title or sheet
    _st, sections = _segment(anchors, covered, nrows, ncols)
    chunks: List[Dict[str, Any]] = []
    for sec in sections:
        sec_title = sec["title"]
        header = sec["header"]
        header_group = sec.get("header_group") or {}
        header_band = sec.get("header_band")
        rows = sec["rows"]
        if not rows:
            continue
        path = [title] + ([sec_title] if sec_title else [])
        no_header = header is None

        def _in_band(c):
            return header_band is None or (header_band[0] <= c <= header_band[1])

        def _key(c):
            leaf = header.get(c) if header else None
            if not leaf:
                return get_column_letter(c)
            g = header_group.get(c)
            return f"{g}_{leaf}" if g and g != leaf else leaf

        def emit_row(cells, orow, rng, source, review, conf, warns):
            if header:
                fields = {_key(c): clean_val(t) for (c, t, cs, rs, tag) in cells if t.strip() and _in_band(c)}
            else:
                fields = {get_column_letter(c): clean_val(t) for (c, t, cs, rs, tag) in cells if t.strip()}
            if not fields:
                return
            fld_txt = ", ".join(f"{k}={v}" for k, v in fields.items())
            content = f"{title}의 {sheet} 시트"
            if sec_title:
                content += f" '{sec_title}' 섹션"
            content += f"에서 다음 값을 가진다: {fld_txt}."
            core = (f"title: {title}; path: {' > '.join(path)}; "
                    + "; ".join(f"{k}: {v}" for k, v in fields.items()) + f" -- {sheet} [{rng}]")
            chunks.append({
                "id": f"{doc}::{sheet}::{rng}::row", "source_file": doc, "sheet": sheet, "range": rng,
                "chunk_type": "table_row", "region_type": ("unknown_table" if no_header else "flat_table"),
                "title": title, "path": path, "fields": fields,
                "facts": [{"predicate": k, "value": v} for k, v in fields.items()],
                "content_text": content[:600], "keywords": _kw(title, sec_title, *fields.values()),
                "source": source,
                "metadata": {"workbook_title": title, "section": sec_title, "core_text": core[:900], "embedding_text": core[:900]},
                "quality": _quality(conf, review, warns),
            })

        for (r, cells) in rows:
            orow = k2o.get(r)
            cols = [c for (c, t, cs, rs, tag) in cells if t.strip()]
            if not cols:
                continue
            c1, c2 = min(cols), max(cols)
            rng = f"{get_column_letter(c1)}{orow}:{get_column_letter(c2)}{orow}" if orow else None
            source = {"file": doc, "sheet": sheet, "range": rng, "start_row": orow,
                      "end_row": orow, "start_col": c1, "end_col": c2}
            warns = (["header_not_detected"] if no_header else []) + ([] if orow else ["coord_unresolved"])
            review = no_header or (orow is None)
            conf = 0.86 if (not no_header and orow) else 0.6

            marker_cells = [(c, t.strip()) for (c, t, cs, rs, tag) in cells if is_marker(t) and _in_band(c)]
            text_cells = [(c, t) for (c, t, cs, rs, tag) in cells if t.strip() and not is_marker(t) and _in_band(c)]
            if header and marker_cells:
                primary_label = header.get(min(header)) if header else None
                desc = {(header.get(c) or primary_label or get_column_letter(c)): clean_val(t) for (c, t) in text_cells}
                row_label = " > ".join(desc.values()) if desc else f"row{orow}"
                buckets: Dict[str, List[str]] = {}
                for (c, t) in marker_cells:
                    bk = MARK_BUCKET.get(t, "해당")
                    buckets.setdefault(bk, []).append(header.get(c) or get_column_letter(c))
                fields = dict(desc)
                for bk, cl in buckets.items():
                    fields[bk] = ", ".join(cl)
                grp = "; ".join(f"{bk}: {', '.join(cl)}" for bk, cl in buckets.items())
                content = f"{title}의 {sheet} 시트에서 '{row_label}' 항목 — {grp}."
                core = (f"title: {title}; path: {' > '.join(path)}; "
                        + "; ".join(f"{k}: {v}" for k, v in fields.items()) + f" -- {sheet} [{rng}]")
                facts = [{"subject": row_label, "predicate": bk, "object": col}
                         for bk, cl in buckets.items() for col in cl]
                chunks.append({
                    "id": f"{doc}::{sheet}::{rng}::matrix", "source_file": doc, "sheet": sheet, "range": rng,
                    "chunk_type": "matrix_fact", "region_type": "matrix_table", "title": title,
                    "path": path + [row_label], "fields": fields, "facts": facts,
                    "content_text": content[:600],
                    "keywords": _kw(title, sec_title, row_label, *[v for cl in buckets.values() for v in cl]),
                    "source": source,
                    "metadata": {"workbook_title": title, "section": sec_title, "core_text": core[:900], "embedding_text": core[:900]},
                    "quality": _quality(conf, review, warns),
                })
            else:
                emit_row(cells, orow, rng, source, review, conf, warns)
    return chunks


# ─── md 확보 ─────────────────────────────────────────
def _resolve_md(input_path: Path, config: ParserConfig) -> str:
    stem = input_path.stem
    if config.kordoc_md_path and Path(config.kordoc_md_path).exists():
        return Path(config.kordoc_md_path).read_text(encoding="utf-8")
    if config.kordoc_md_dir:
        p = Path(config.kordoc_md_dir) / f"{stem}.md"
        if p.exists():
            return p.read_text(encoding="utf-8")
    # 자동 생성 (Node 필요)
    if config.kordoc_bin:
        out_dir = Path(config.kordoc_md_out or tempfile.gettempdir())
        out_dir.mkdir(parents=True, exist_ok=True)
        md_path = out_dir / f"{stem}.md"
        # kordoc CLI: `kordoc <file> -o <out.md> --silent`. kordoc_bin 은 멀티워드 허용
        # (예: "node /path/dist/cli.js").
        cmd = shlex.split(config.kordoc_bin) + [str(input_path), "-o", str(md_path), "--silent"]
        proc = subprocess.run(cmd, capture_output=True, text=True)
        if proc.returncode != 0 or not md_path.exists():
            raise BackendError(
                f"kordoc 자동생성 실패 (bin={config.kordoc_bin}): {proc.stderr[:300] or proc.stdout[:300]}")
        return md_path.read_text(encoding="utf-8")
    raise BackendError(
        f"kordoc backend: '{stem}.md' 를 찾을 수 없습니다. "
        f"--kordoc-md/--kordoc-md-dir 로 지정하거나 --kordoc-bin 으로 자동생성하세요.")


class KordocBackend:
    def parse(self, input_path: Path, config: ParserConfig) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
        md = _resolve_md(input_path, config)
        doc = input_path.stem
        wb = load_workbook(input_path, data_only=True)
        ksheets = dict(_split_sheets(md))
        all_chunks: List[Dict[str, Any]] = []
        sheets_done = 0
        for sheet in wb.sheetnames:
            body = ksheets.get(sheet)
            if body is None:
                continue
            anchors, covered, nrows, kncols = _grid(body)
            ws = wb[sheet]
            acols = max(kncols, ws.max_column)
            k2o = _align_rows(anchors, covered, nrows, acols, ws)
            sheet_title, _secs = _segment(anchors, covered, nrows, kncols)
            all_chunks.extend(_build(doc, sheet, sheet_title, anchors, covered, nrows, kncols, k2o))
            sheets_done += 1

        ct = Counter(c["chunk_type"] for c in all_chunks)
        rt = Counter(c["region_type"] for c in all_chunks)
        confs = [c["quality"]["confidence"] for c in all_chunks] or [0.0]
        stats = {
            "backend": "kordoc",
            "source_file": input_path.name,
            "sheet_count": len(wb.sheetnames),
            "sheets_parsed": sheets_done,
            "total_chunks": len(all_chunks),
            "chunk_type_counts": dict(ct),
            "region_type_counts": dict(rt),
            "coords": sum(1 for c in all_chunks if c["range"]),
            "confidence": {"avg": round(sum(confs) / len(confs), 4), "min": min(confs), "max": max(confs)},
            "review_required_count": sum(1 for c in all_chunks if c["quality"]["review_required"]),
        }
        return all_chunks, stats
