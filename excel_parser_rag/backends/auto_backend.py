"""AutoBackend — 전결 문서만 openpyxl 로 라우팅하는 2-tier 라우터 (spec 2026-06-26).

Tier1: 헤더 영역에 "전결" 토큰이 있으면 후보 → openpyxl 시도.
Tier2: openpyxl 결과에 delegation_rule 이 1개 이상이면 채택, 아니면 kordoc fallback.
키워드 없으면 즉시 kordoc (openpyxl 미접촉 → 평면목록 오분류 원천 차단).
"""
from __future__ import annotations

import copy
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl

from ..config import ParserConfig
from .base import Backend

_DELEGATION_TOKEN = "전결"
_OPENPYXL_PROFILES = ["delegation_rule", "note", "code_mapping"]
# openpyxl 직접 읽기가 가능한 확장자만 라우팅 대상. .xls 는 openpyxl.load_workbook 로 못 열고
# (소속 backend 가 soffice 변환을 내부 처리), 변환 비용/복잡도가 커서 kordoc 로 보낸다(현행 동일).
_OPENPYXL_SUFFIXES = {".xlsx", ".xlsm"}


def detect_delegation_keyword(input_path, *, max_rows: int = 40, max_cols: int = 30) -> bool:
    """워크북 상단 영역에 "전결" 토큰이 한 번이라도 나오면 True."""
    path = Path(input_path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
                for v in row:
                    if isinstance(v, str) and _DELEGATION_TOKEN in v:
                        return True
        return False
    finally:
        wb.close()


def _should_try_openpyxl(input_path) -> bool:
    """Tier1: openpyxl 가능 확장자(.xlsx/.xlsm) + "전결" 키워드일 때만 True.
    `and` 단락 평가로 .xls 는 detect 를 호출하지 않는다(파일 미존재여도 안전)."""
    return (
        Path(input_path).suffix.lower() in _OPENPYXL_SUFFIXES
        and detect_delegation_keyword(input_path)
    )


def _with(config: ParserConfig, backend: str, *, profiles=None) -> ParserConfig:
    cfg = copy.copy(config)
    cfg.backend = backend
    if profiles is not None:
        cfg.chunk_profiles = list(profiles)
    return cfg


class AutoBackend(Backend):
    name = "auto"

    def parse(self, input_path, config: ParserConfig) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
        from . import get_backend  # 지연 import (순환 방지)

        def _kordoc():
            chunks, stats = get_backend("kordoc").parse(input_path, _with(config, "kordoc"))
            stats = dict(stats); stats["routed_backend"] = "kordoc"
            return chunks, stats

        # Tier1: .xls 등 비대상 확장자 또는 "전결" 키워드 없음 → kordoc (현행 동작 보존)
        if not _should_try_openpyxl(input_path):
            return _kordoc()

        op_cfg = _with(config, "openpyxl", profiles=_OPENPYXL_PROFILES)
        chunks, stats = get_backend("openpyxl").parse(input_path, op_cfg)
        if any(c.get("chunk_type") == "delegation_rule" for c in chunks):
            stats = dict(stats); stats["routed_backend"] = "openpyxl"
            return chunks, stats
        # 키워드만 우연히 박힌 문서 → kordoc fallback
        return _kordoc()
