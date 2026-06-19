"""Worksheet → SheetCanvas 변환 (SoT §4 step 3, §6.1~6.2).

pipeline.py 가 고정한 시그니처:

- build_sheet_canvas(ws, formula_ws, workbook_name, sheet_index, config) -> SheetCanvas

used range 의 모든 셀을 CellNode 로 변환해 canvas 에 넣는다.
빈 셀도 스타일(테두리/배경)이 region/header 탐지의 신호이므로 보존한다.
"""

from __future__ import annotations

import datetime
from typing import Any, Optional, Tuple

from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from ..canvas.cell_node import CellNode, CellStyle
from ..canvas.sheet_canvas import SheetCanvas
from ..config import ParserConfig
from ..textutil import clean_text, one_line

# 사실상 "배경 없음"으로 취급하는 RGB 값
_NO_FILL_RGB = {"FFFFFFFF", "00000000", "FFFFFF", "000000"}


def _color_to_hex(color: Any) -> Optional[str]:
    """openpyxl Color → 'FFRRGGBB' hex 문자열. theme/indexed 색은 None.

    fgColor 가 theme 타입이면 .rgb 가 예외성 값(없거나 int)일 수 있으므로
    type == 'rgb' 이고 .rgb 가 str 일 때만 신뢰한다.
    """
    if color is None:
        return None
    try:
        if getattr(color, "type", None) != "rgb":
            return None
        rgb = color.rgb
    except (AttributeError, TypeError):
        return None
    if isinstance(rgb, str) and rgb:
        return rgb
    return None


def _extract_fill_color(cell: Any) -> Optional[str]:
    """배경색 RGB hex. 패턴 없음/theme/흰색·투명은 None."""
    fill = cell.fill
    if fill is None or getattr(fill, "patternType", None) is None:
        return None
    rgb = _color_to_hex(getattr(fill, "fgColor", None))
    if rgb is None:
        rgb = _color_to_hex(getattr(fill, "start_color", None))
    if rgb is None or rgb.upper() in _NO_FILL_RGB:
        return None
    return rgb


def _extract_style(cell: Any) -> CellStyle:
    font = cell.font
    alignment = cell.alignment
    border = cell.border

    font_size: Optional[float] = None
    if font is not None and font.size is not None:
        try:
            font_size = float(font.size)
        except (TypeError, ValueError):
            font_size = None

    indent = 0
    if alignment is not None and alignment.indent:
        try:
            indent = int(alignment.indent)
        except (TypeError, ValueError):
            indent = 0

    def _has_side(side: Any) -> bool:
        return side is not None and side.style is not None

    return CellStyle(
        bold=bool(font.bold) if font is not None else False,
        italic=bool(font.italic) if font is not None else False,
        font_size=font_size,
        fill_color=_extract_fill_color(cell),
        font_color=_color_to_hex(font.color) if font is not None else None,
        horizontal_alignment=alignment.horizontal if alignment is not None else None,
        vertical_alignment=alignment.vertical if alignment is not None else None,
        border_top=_has_side(border.top) if border is not None else False,
        border_bottom=_has_side(border.bottom) if border is not None else False,
        border_left=_has_side(border.left) if border is not None else False,
        border_right=_has_side(border.right) if border is not None else False,
        number_format=cell.number_format,
        indent=indent,
    )


def _format_value(value: Any) -> str:
    """raw value → 사람이 읽는 표시 문자열 (정규화 전)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime.datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, datetime.time):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def _detect_data_type(value: Any, formula_text: Optional[str]) -> str:
    """empty|str|int|float|date|bool|formula 판정 (SoT §6.1)."""
    if value is None:
        # 캐시값이 없는 수식 셀은 formula 로 남긴다
        return "formula" if formula_text else "empty"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return "date"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    return "str"


def _extract_formula_text(formula_ws: Optional[Worksheet], row: int, col: int) -> Optional[str]:
    if formula_ws is None:
        return None
    try:
        value = formula_ws.cell(row=row, column=col).value
    except (IndexError, ValueError):
        return None
    if isinstance(value, str) and value.startswith("="):
        return value
    # ArrayFormula 등 객체형 수식
    text = getattr(value, "text", None)
    if isinstance(text, str) and text.startswith("="):
        return text
    return None


def _used_range(ws: Worksheet, config: ParserConfig) -> Tuple[int, int, bool]:
    """(유효 max_row, 유효 max_col, 잘림 여부)."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row <= 0 or max_col <= 0:
        return 0, 0, False
    truncated = False
    limit = config.max_cells_per_sheet
    if limit and max_row * max_col > limit:
        max_row = max(1, limit // max_col)
        truncated = True
    return max_row, max_col, truncated


def build_sheet_canvas(
    ws: Worksheet,
    formula_ws: Optional[Worksheet],
    workbook_name: str,
    sheet_index: int,
    config: ParserConfig,
) -> SheetCanvas:
    """openpyxl worksheet → SheetCanvas (스타일/병합/숨김/수식 보존)."""
    canvas = SheetCanvas(
        workbook_name=workbook_name,
        sheet_name=ws.title,
        sheet_index=sheet_index,
        is_hidden_sheet=getattr(ws, "sheet_state", "visible") != "visible",
    )

    # 숨김 행/열 — CellNode 생성 전에 채워야 hidden 플래그가 정확하다
    for row_idx, dim in ws.row_dimensions.items():
        if dim.hidden:
            canvas.hidden_rows.add(int(row_idx))
    for col_key, dim in ws.column_dimensions.items():
        if dim.hidden:
            try:
                canvas.hidden_cols.add(column_index_from_string(col_key))
            except ValueError:
                continue

    canvas.merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]

    max_row, max_col, truncated = _used_range(ws, config)
    if truncated:
        # SheetCanvas 에 warnings 필드는 없으므로 동적 속성으로 표시 (디버그/리포트용)
        setattr(canvas, "truncated", True)
        setattr(canvas, "truncated_at_row", max_row)
    if max_row == 0 or max_col == 0:
        return canvas

    for excel_row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in excel_row:
            row, col = cell.row, cell.column
            raw_value = cell.value
            formula_text = _extract_formula_text(formula_ws, row, col)
            if formula_text:
                canvas.contains_formula = True

            display = clean_text(_format_value(raw_value))
            node = CellNode(
                sheet=ws.title,
                row=row,
                col=col,
                raw_value=raw_value,
                display_value=display,
                normalized_value=one_line(display),
                formula=formula_text,
                data_type=_detect_data_type(raw_value, formula_text),
                is_empty=(raw_value is None or display == ""),
                style=_extract_style(cell),
                hidden_row=row in canvas.hidden_rows,
                hidden_col=col in canvas.hidden_cols,
            )
            canvas.put_cell(node)

    # used range 가 dims 보다 작아도 시트 크기는 보존
    canvas.max_row = max(canvas.max_row, max_row)
    canvas.max_col = max(canvas.max_col, max_col)
    return canvas
